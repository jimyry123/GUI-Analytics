from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button, RadioButtons, CheckButtons
from matplotlib.patches import Rectangle
import matplotlib.patches as mpatches
from matplotlib.ticker import PercentFormatter
import getpass

import matplotlib

from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt4agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

competitor = []
year = []
        
class First(QDialog):
    def __init__(self):
        super(First, self).__init__()
        self.createFormGroupBox()
        
    def createFormGroupBox(self):  
        
        self.setGeometry(350,150,300,200)
        
        font = QFont()
        font.setPointSize(20)
        
        vis = QHBoxLayout()
        amat = QHBoxLayout()
        competitors = QHBoxLayout()
        
        buttonLayout1 = QVBoxLayout()
        
        self.competitors = QPushButton("Input Quarters")
        
        self.amat = QPushButton("AMAT")
        self.visual = QPushButton("Perform visualization")
        
        self.setFont(font)
        
        competitors.addWidget(self.competitors)
        amat.addWidget(self.amat)
        vis.addWidget(self.visual)
        buttonLayout1.addLayout(competitors)
        buttonLayout1.addLayout(amat)
        buttonLayout1.addLayout(vis)
        mainLayout = QGridLayout()
        # mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addLayout(buttonLayout1, 0, 1)
        self.setLayout(mainLayout)
        self.setLayout(buttonLayout1)
        self.setWindowTitle("Main Page")
        self.amat.clicked.connect(self.on_pushButton_clicked_1)
        self.competitors.clicked.connect(self.on_pushButton_clicked)
        self.visual.clicked.connect(self.on_pushButton_clicked_2)
        self.dialogs = list()
        
        self.show()

    def on_pushButton_clicked(self):
        dialog = Second(self)
        self.dialogs.append(dialog)
        dialog.show()
        
    def on_pushButton_clicked_1(self):
        dialog = Third(self)
        self.dialogs.append(dialog)
        dialog.show()
    
    def on_pushButton_clicked_2(self):
        dialog = Fourth(self)
        self.dialogs.append(dialog)
        dialog.show()

class Fourth(QDialog):
    def __init__(self, parent = None):
        super(Fourth, self).__init__(parent)
        self.createFormGroupBox()
    
    def createFormGroupBox(self):
        year = QHBoxLayout()
        quarter = QHBoxLayout()
        features = QHBoxLayout()
        graph2 = QHBoxLayout()
        
        buttonLayout1 = QVBoxLayout()
        
        self.year = QPushButton("View Equipment over competitor")
        
        self.quarter = QPushButton("View and Download table")
        self.features = QPushButton("View shares over competitor")
            
        year.addWidget(self.year)
        features.addWidget(self.features)
        quarter.addWidget(self.quarter)
        buttonLayout1.addLayout(year)
        buttonLayout1.addLayout(features)
        buttonLayout1.addLayout(quarter)
        mainLayout = QGridLayout()
        # mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addLayout(buttonLayout1, 0, 1)
        self.setLayout(mainLayout)
        self.setLayout(buttonLayout1)
        self.setWindowTitle("Visualization")
        self.year.clicked.connect(self.on_pushButton_clicked)
        self.features.clicked.connect(self.on_pushButton_clicked_2)
        self.quarter.clicked.connect(self.on_pushButton_clicked_1)
        self.dialogs = list()
        
        self.show()
        
    def on_pushButton_clicked(self):
        dialog = Window(self)
        self.dialogs.append(dialog)
        dialog.show()

    def on_pushButton_clicked_1(self):
        dialog = Export(self)
        self.dialogs.append(dialog)
        dialog.show()
    
    def on_pushButton_clicked_2(self):
        dialog = line(self)
        self.dialogs.append(dialog)
        dialog.show()

class App(QDialog):
    def __init__(self, parent=None):
        super(App, self).__init__(parent)
        self.title = 'Tables'
        self.left = 100
        self.top = 50
        self.width = 700
        self.height = 400
        self.initUI()
 
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        
        for items in year:
            if items == '2018':
                self.createTable()
            if items == '2017':
                self.createTable1()
            if items == '2016':
                self.table2016()
            if items == '2015':
                self.table2015()
            if items =='2014':
                self.table2014()
            if items == '2013':
                self.table2013()
        
        self.buttonSave = QPushButton('Download', self)
        self.buttonSave.clicked.connect(self.savefile)
        
        # Add box layout, add table to box layout and add box layout to widget
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.tableWidget) 
        self.layout.addWidget(self.buttonSave)
        self.setLayout(self.layout) 
 
        # Show widget
        self.show()

    def createTable(self):
       # Create table
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
        
        nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY18')]
        
        nlpcvd_num = (nlpcvd_filter['Yearly'])
        
        pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY18')]
        
        pcvd_num= (pcvd_filter['Yearly'])
        
        aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY18')]
        
        aldp_num= (aldp_filter['Yearly'])
        
        tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY18')]
        
        tcvd_num= (tcvd_filter['Yearly'])
        
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(14)
        self.tableWidget.setColumnCount(5)
        comp_list = ["Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
        num1 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        num2 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        set_num = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        nlpcvd = [1,1,1,1,1,1,1,1,1,1,1,1,1,1]
        pcvd = [2,2,2,2,2,2,2,2,2,2,2,2,2,2]
        aldp = [3,3,3,3,3,3,3,3,3,3,3,3,3,3]
        tcvd = [4,4,4,4,4,4,4,4,4,4,4,4,4,4]
        
        
        for value1, value2, value3 in zip(num1, set_num, comp_list):
            self.tableWidget.setItem(0,0, QTableWidgetItem("Competitor"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(value3))
        for value1, value2, value3 in zip(num2,nlpcvd, nlpcvd_num):
            self.tableWidget.setItem(0,1, QTableWidgetItem("NLPCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,pcvd, pcvd_num):
            self.tableWidget.setItem(0,2, QTableWidgetItem("PCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,aldp, aldp_num):
            self.tableWidget.setItem(0,3, QTableWidgetItem("ALDP"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,tcvd, tcvd_num):
            self.tableWidget.setItem(0,4, QTableWidgetItem("TCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
            
        
    def createTable1(self):
       # Create table
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
        
        nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY17')]
        
        nlpcvd_num = (nlpcvd_filter['Yearly'])
        
        pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY17')]
        
        pcvd_num= (pcvd_filter['Yearly'])
        
        aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY17')]
        
        aldp_num= (aldp_filter['Yearly'])
        
        tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY17')]
        
        tcvd_num= (tcvd_filter['Yearly'])
        
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(14)
        self.tableWidget.setColumnCount(5)
        comp_list = ["Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
        num1 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        num2 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        set_num = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        nlpcvd = [1,1,1,1,1,1,1,1,1,1,1,1,1,1]
        pcvd = [2,2,2,2,2,2,2,2,2,2,2,2,2,2]
        aldp = [3,3,3,3,3,3,3,3,3,3,3,3,3,3]
        tcvd = [4,4,4,4,4,4,4,4,4,4,4,4,4,4]
        
        
        for value1, value2, value3 in zip(num1, set_num, comp_list):
            self.tableWidget.setItem(0,0, QTableWidgetItem("Competitor"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(value3))
        for value1, value2, value3 in zip(num2,nlpcvd, nlpcvd_num):
            self.tableWidget.setItem(0,1, QTableWidgetItem("NLPCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,pcvd, pcvd_num):
            self.tableWidget.setItem(0,2, QTableWidgetItem("PCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,aldp, aldp_num):
            self.tableWidget.setItem(0,3, QTableWidgetItem("ALDP"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,tcvd, tcvd_num):
            self.tableWidget.setItem(0,4, QTableWidgetItem("TCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
            
    def table2016(self):
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
        
        nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY16')]
        
        nlpcvd_num = (nlpcvd_filter['Yearly'])
        
        pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY16')]
        
        pcvd_num= (pcvd_filter['Yearly'])
        
        aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY16')]
        
        aldp_num= (aldp_filter['Yearly'])
        
        tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY16')]
        
        tcvd_num= (tcvd_filter['Yearly'])
        
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(14)
        self.tableWidget.setColumnCount(5)
        comp_list = [ "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
        num1 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        num2 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        set_num = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        nlpcvd = [1,1,1,1,1,1,1,1,1,1,1,1,1,1]
        pcvd = [2,2,2,2,2,2,2,2,2,2,2,2,2,2]
        aldp = [3,3,3,3,3,3,3,3,3,3,3,3,3,3]
        tcvd = [4,4,4,4,4,4,4,4,4,4,4,4,4,4]
        
        
        for value1, value2, value3 in zip(num1, set_num, comp_list):
            self.tableWidget.setItem(0,0, QTableWidgetItem("Competitor"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(value3))
        for value1, value2, value3 in zip(num2,nlpcvd, nlpcvd_num):
            self.tableWidget.setItem(0,1, QTableWidgetItem("NLPCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,pcvd, pcvd_num):
            self.tableWidget.setItem(0,2, QTableWidgetItem("PCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,aldp, aldp_num):
            self.tableWidget.setItem(0,3, QTableWidgetItem("ALDP"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,tcvd, tcvd_num):
            self.tableWidget.setItem(0,4, QTableWidgetItem("TCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
            
    def table2015(self):
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
        
        nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY15')]
        
        nlpcvd_num = (nlpcvd_filter['Yearly'])
        
        pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY15')]
        
        pcvd_num= (pcvd_filter['Yearly'])
        
        aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY15')]
        
        aldp_num= (aldp_filter['Yearly'])
        
        tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY15')]
        
        tcvd_num= (tcvd_filter['Yearly'])
        
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(14)
        self.tableWidget.setColumnCount(5)
        comp_list = ["Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
        num1 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        num2 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        set_num = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        nlpcvd = [1,1,1,1,1,1,1,1,1,1,1,1,1,1]
        pcvd = [2,2,2,2,2,2,2,2,2,2,2,2,2,2]
        aldp = [3,3,3,3,3,3,3,3,3,3,3,3,3,3]
        tcvd = [4,4,4,4,4,4,4,4,4,4,4,4,4,4]
        
        
        for value1, value2, value3 in zip(num1, set_num, comp_list):
            self.tableWidget.setItem(0,0, QTableWidgetItem("Competitor"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(value3))
        for value1, value2, value3 in zip(num2,nlpcvd, nlpcvd_num):
            self.tableWidget.setItem(0,1, QTableWidgetItem("NLPCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,pcvd, pcvd_num):
            self.tableWidget.setItem(0,2, QTableWidgetItem("PCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,aldp, aldp_num):
            self.tableWidget.setItem(0,3, QTableWidgetItem("ALDP"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,tcvd, tcvd_num):
            self.tableWidget.setItem(0,4, QTableWidgetItem("TCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
    
    def table2014(self):
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
        
        nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY14')]
        
        nlpcvd_num = (nlpcvd_filter['Yearly'])
        
        pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY14')]
        
        pcvd_num= (pcvd_filter['Yearly'])
        
        aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY14')]
        
        aldp_num= (aldp_filter['Yearly'])
        
        tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY14')]
        
        tcvd_num= (tcvd_filter['Yearly'])
        
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(14)
        self.tableWidget.setColumnCount(5)
        comp_list = ["Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
        num1 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        num2 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        set_num = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        nlpcvd = [1,1,1,1,1,1,1,1,1,1,1,1,1,1]
        pcvd = [2,2,2,2,2,2,2,2,2,2,2,2,2,2]
        aldp = [3,3,3,3,3,3,3,3,3,3,3,3,3,3]
        tcvd = [4,4,4,4,4,4,4,4,4,4,4,4,4,4]
        
        
        for value1, value2, value3 in zip(num1, set_num, comp_list):
            self.tableWidget.setItem(0,0, QTableWidgetItem("Competitor"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(value3))
        for value1, value2, value3 in zip(num2,nlpcvd, nlpcvd_num):
            self.tableWidget.setItem(0,1, QTableWidgetItem("NLPCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,pcvd, pcvd_num):
            self.tableWidget.setItem(0,2, QTableWidgetItem("PCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,aldp, aldp_num):
            self.tableWidget.setItem(0,3, QTableWidgetItem("ALDP"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,tcvd, tcvd_num):
            self.tableWidget.setItem(0,4, QTableWidgetItem("TCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
            
    def table2013(self):
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
        
        nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY13')]
        
        nlpcvd_num = (nlpcvd_filter['Yearly'])
        
        pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY13')]
        
        pcvd_num= (pcvd_filter['Yearly'])
        
        aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY13')]
        
        aldp_num= (aldp_filter['Yearly'])
        
        tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY13')]
        
        tcvd_num= (tcvd_filter['Yearly'])
        
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(14)
        self.tableWidget.setColumnCount(5)
        comp_list = ["Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
        num1 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        num2 = [1,2,3,4,5,6,7,8,9,10,11,12,13]
        set_num = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        nlpcvd = [1,1,1,1,1,1,1,1,1,1,1,1,1,1]
        pcvd = [2,2,2,2,2,2,2,2,2,2,2,2,2,2]
        aldp = [3,3,3,3,3,3,3,3,3,3,3,3,3,3]
        tcvd = [4,4,4,4,4,4,4,4,4,4,4,4,4,4]
        
        
        for value1, value2, value3 in zip(num1, set_num, comp_list):
            self.tableWidget.setItem(0,0, QTableWidgetItem("Competitor"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(value3))
        for value1, value2, value3 in zip(num2,nlpcvd, nlpcvd_num):
            self.tableWidget.setItem(0,1, QTableWidgetItem("NLPCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,pcvd, pcvd_num):
            self.tableWidget.setItem(0,2, QTableWidgetItem("PCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,aldp, aldp_num):
            self.tableWidget.setItem(0,3, QTableWidgetItem("ALDP"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        for value1, value2, value3 in zip(num2,tcvd, tcvd_num):
            self.tableWidget.setItem(0,4, QTableWidgetItem("TCVD"))
            self.tableWidget.setItem(value1,value2, QTableWidgetItem(("{}").format(value3)))
        
    def savefile(self):
        filename = QFileDialog.getSaveFileName(self, "Save to XLSX", ".xlsx",
                                                       "Comma Separated Values Spreadsheet (*.xlsx);;"
                                                       "All Files (*)")[0]
        for items in year:
            if items == '2018':
                
                value = [1,2,3,4,5,6,7,8,9,10,11,12,13,14]
                new_value = [2,3,4,5,6,7,8,9,10,11,12,13,14]
                nlpcvd = []
                pcvd = []
                aldp = []
                tcvd = []
                
                comp = ["Competitor", "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
                new = []
                
                wb = Workbook()
                ws1 = wb.active
                
                x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
                x1.sheet_names
                df = x1.parse("Revised")
                
                df['Yearly'].fillna(0, inplace = True)
                
                df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
                
                nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY18')]
                
                nlpcvd_num = (nlpcvd_filter['Yearly'])
                
                pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY18')]
                
                pcvd_num= (pcvd_filter['Yearly'])
                
                aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY18')]
                
                aldp_num= (aldp_filter['Yearly'])
                
                tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY18')]
                
                tcvd_num= (tcvd_filter['Yearly'])
                
                for num in value:
                    new.append("A{}".format(num))
                    
                for num in new_value:
                    nlpcvd.append("B{}".format(num))
                
                for num in new_value:
                    pcvd.append("C{}".format(num))
                
                for num in new_value:
                    aldp.append("D{}".format(num))
                    
                for num in new_value:
                    tcvd.append("E{}".format(num))
                    
                for num, values in zip(new, comp):
                    ws1[num] =values
                    ws1['B1'] = 'NLPCVD'
                    ws1['C1'] = 'PCVD'
                    ws1['D1'] = 'ALDP'
                    ws1['E1'] = 'TCVD'
                    
                    wb.save(filename)
                
                for num, values in zip(nlpcvd, nlpcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(pcvd, pcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(aldp, aldp_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(tcvd, tcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                
                
        for items in year:
            if items == '2017':
                
                value = [1,2,3,4,5,6,7,8,9,10,11,12,13,14]
                new_value = [2,3,4,5,6,7,8,9,10,11,12,13,14]
                nlpcvd = []
                pcvd = []
                aldp = []
                tcvd = []
                
                comp = ["Competitor", "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
                new = []
                
                wb = Workbook()
                ws1 = wb.active
                
                x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
                x1.sheet_names
                df = x1.parse("Revised")
                
                df['Yearly'].fillna(0, inplace = True)
                
                df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
                
                nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY17')]
                
                nlpcvd_num = (nlpcvd_filter['Yearly'])
                
                pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY17')]
                
                pcvd_num= (pcvd_filter['Yearly'])
                
                aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY17')]
                
                aldp_num= (aldp_filter['Yearly'])
                
                tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY17')]
                
                tcvd_num= (tcvd_filter['Yearly'])
                
                for num in value:
                    new.append("A{}".format(num))
                    
                for num in new_value:
                    nlpcvd.append("B{}".format(num))
                
                for num in new_value:
                    pcvd.append("C{}".format(num))
                
                for num in new_value:
                    aldp.append("D{}".format(num))
                    
                for num in new_value:
                    tcvd.append("E{}".format(num))
                    
                for num, values in zip(new, comp):
                    ws1[num] =values
                    ws1['B1'] = 'NLPCVD'
                    ws1['C1'] = 'PCVD'
                    ws1['D1'] = 'ALDP'
                    ws1['E1'] = 'TCVD'
                    
                    wb.save(filename)
                
                for num, values in zip(nlpcvd, nlpcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(pcvd, pcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(aldp, aldp_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(tcvd, tcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                
        for items in year:
            if items == '2016':
                
                value = [1,2,3,4,5,6,7,8,9,10,11,12,13,14]
                new_value = [2,3,4,5,6,7,8,9,10,11,12,13,14]
                nlpcvd = []
                pcvd = []
                aldp = []
                tcvd = []
                
                comp = ["Competitor", "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
                new = []
                
                wb = Workbook()
                ws1 = wb.active
                
                x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
                x1.sheet_names
                df = x1.parse("Revised")
                
                df['Yearly'].fillna(0, inplace = True)
                
                df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
                
                nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY16')]
                
                nlpcvd_num = (nlpcvd_filter['Yearly'])
                
                pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY16')]
                
                pcvd_num= (pcvd_filter['Yearly'])
                
                aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY16')]
            
                aldp_num= (aldp_filter['Yearly'])
                
                tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY16')]
                
                tcvd_num= (tcvd_filter['Yearly'])
                
                for num in value:
                    new.append("A{}".format(num))
                    
                for num in new_value:
                    nlpcvd.append("B{}".format(num))
                
                for num in new_value:
                    pcvd.append("C{}".format(num))
                
                for num in new_value:
                    aldp.append("D{}".format(num))
                    
                for num in new_value:
                    tcvd.append("E{}".format(num))
                    
                for num, values in zip(new, comp):
                    ws1[num] =values
                    ws1['B1'] = 'NLPCVD'
                    ws1['C1'] = 'PCVD'
                    ws1['D1'] = 'ALDP'
                    ws1['E1'] = 'TCVD'
                    
                    wb.save(filename)
                
                for num, values in zip(nlpcvd, nlpcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(pcvd, pcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(aldp, aldp_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(tcvd, tcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                
                
        for items in year:
            if items == '2015':
                
                value = [1,2,3,4,5,6,7,8,9,10,11,12,13,14]
                new_value = [2,3,4,5,6,7,8,9,10,11,12,13,14]
                nlpcvd = []
                pcvd = []
                aldp = []
                tcvd = []
                
                comp = ["Competitor", "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
                new = []
                
                wb = Workbook()
                ws1 = wb.active
                
                x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
                x1.sheet_names
                df = x1.parse("Revised")
                
                df['Yearly'].fillna(0, inplace = True)
                
                df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
                
                nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY15')]
                
                nlpcvd_num = (nlpcvd_filter['Yearly'])
                
                pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY15')]
                
                pcvd_num= (pcvd_filter['Yearly'])
                
                aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY15')]
            
                aldp_num= (aldp_filter['Yearly'])
                
                tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY15')]
                
                tcvd_num= (tcvd_filter['Yearly'])
                
                for num in value:
                    new.append("A{}".format(num))
                    
                for num in new_value:
                    nlpcvd.append("B{}".format(num))
                
                for num in new_value:
                    pcvd.append("C{}".format(num))
                
                for num in new_value:
                    aldp.append("D{}".format(num))
                    
                for num in new_value:
                    tcvd.append("E{}".format(num))
                    
                for num, values in zip(new, comp):
                    ws1[num] =values
                    ws1['B1'] = 'NLPCVD'
                    ws1['C1'] = 'PCVD'
                    ws1['D1'] = 'ALDP'
                    ws1['E1'] = 'TCVD'
                    
                    wb.save(filename)
                
                for num, values in zip(nlpcvd, nlpcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(pcvd, pcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(aldp, aldp_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(tcvd, tcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                
        for items in year:
            if items == '2014':
                
                value = [1,2,3,4,5,6,7,8,9,10,11,12,13,14]
                new_value = [2,3,4,5,6,7,8,9,10,11,12,13,14]
                nlpcvd = []
                pcvd = []
                aldp = []
                tcvd = []
                
                comp = ["Competitor", "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
                new = []
                
                wb = Workbook()
                ws1 = wb.active
                
                x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
                x1.sheet_names
                df = x1.parse("Revised")
                
                df['Yearly'].fillna(0, inplace = True)
                
                df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
                
                nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY14')]
                
                nlpcvd_num = (nlpcvd_filter['Yearly'])
                
                pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY14')]
                
                pcvd_num= (pcvd_filter['Yearly'])
                
                aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY14')]
            
                aldp_num= (aldp_filter['Yearly'])
                
                tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY14')]
                
                tcvd_num= (tcvd_filter['Yearly'])
                
                for num in value:
                    new.append("A{}".format(num))
                    
                for num in new_value:
                    nlpcvd.append("B{}".format(num))
                
                for num in new_value:
                    pcvd.append("C{}".format(num))
                
                for num in new_value:
                    aldp.append("D{}".format(num))
                    
                for num in new_value:
                    tcvd.append("E{}".format(num))
                    
                for num, values in zip(new, comp):
                    ws1[num] =values
                    ws1['B1'] = 'NLPCVD'
                    ws1['C1'] = 'PCVD'
                    ws1['D1'] = 'ALDP'
                    ws1['E1'] = 'TCVD'
                    
                    wb.save(filename)
                
                for num, values in zip(nlpcvd, nlpcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(pcvd, pcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(aldp, aldp_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(tcvd, tcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                
        for items in year:
            if items == '2013':
                
                value = [1,2,3,4,5,6,7,8,9,10,11,12,13,14]
                new_value = [2,3,4,5,6,7,8,9,10,11,12,13,14]
                nlpcvd = []
                pcvd = []
                aldp = []
                tcvd = []
                
                comp = ["Competitor", "Applied Materials","Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS"]
                new = []
                
                wb = Workbook()
                ws1 = wb.active
                
                x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
                x1.sheet_names
                df = x1.parse("Revised")
                
                df['Yearly'].fillna(0, inplace = True)
                
                df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2', 'Level 1']]
                
                nlpcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Nontube LPCVD') & (df['CY'] == 'CY13')]
                
                nlpcvd_num = (nlpcvd_filter['Yearly'])
                
                pcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Plasma CVD') & (df['CY'] == 'CY13')]
                
                pcvd_num= (pcvd_filter['Yearly'])
                
                aldp_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Atomic Layer Deposition Platforms') & (df['CY'] == 'CY13')]
            
                aldp_num= (aldp_filter['Yearly'])
                
                tcvd_filter = df[(df['Level 1'] == 'Deposition - CVD') & (df['Level 2'] == 'Tube CVD') & (df['CY'] == 'CY13')]
                
                tcvd_num= (tcvd_filter['Yearly'])
                
                for num in value:
                    new.append("A{}".format(num))
                    
                for num in new_value:
                    nlpcvd.append("B{}".format(num))
                
                for num in new_value:
                    pcvd.append("C{}".format(num))
                
                for num in new_value:
                    aldp.append("D{}".format(num))
                    
                for num in new_value:
                    tcvd.append("E{}".format(num))
                    
                for num, values in zip(new, comp):
                    ws1[num] =values
                    ws1['B1'] = 'NLPCVD'
                    ws1['C1'] = 'PCVD'
                    ws1['D1'] = 'ALDP'
                    ws1['E1'] = 'TCVD'
                    
                    wb.save(filename)
                
                for num, values in zip(nlpcvd, nlpcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(pcvd, pcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(aldp, aldp_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                    
                for num, values in zip(tcvd, tcvd_num):
                    ws1[num] =values
                    
                    wb.save(filename)
                
                
        QMessageBox.information(self, "Success",
                "Your download has been completed.")
        
        
class Export(QDialog):
    def __init__(self, parent = None):
        super(Export, self).__init__(parent)
        self.createFormGroupBox()
        
    def createFormGroupBox(self):  
        
        font = QFont()
        font.setPointSize(16)
        buttonLayout1 = QVBoxLayout()
        buttons = QHBoxLayout()
        year = QHBoxLayout()
                     
        self.year = QPushButton("Select the year:")
        self.year.clicked.connect(self.getItem)

        self.nameLine = QLineEdit()
                 
        self.update = QPushButton("View Table")
        self.update.clicked.connect(self.table_13)
        self.cancel = QPushButton("Finish")
        self.cancel.clicked.connect(self.submitContact)
    
        self.setFont(font)
                
        year.addWidget(self.year)
        year.addWidget(self.nameLine)
        buttonLayout1.addLayout(year)
        
        buttons.addWidget(self.update)
        buttons.addWidget(self.cancel)
        buttonLayout1.addLayout(buttons)   
        
        mainLayout = QGridLayout()
        # mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addLayout(buttonLayout1, 0, 1)
 
        self.setLayout(mainLayout)
        self.setLayout(buttonLayout1)
        self.setWindowTitle("Hello Qt")
    
    def getItem(self):
        items1 = ("2018", "2017", "2016", "2015", "2014", "2013")
		
        items, ok = QInputDialog.getItem(self, "select input dialog", 
         "list of competitors", items1, 0, False)
			
        if ok and items:
            self.nameLine.setText(items)
            year.append(items)
            print(items)
    
    def table_13(self):
        year = self.nameLine.text()
        
        if year != '':
            dialog = App(self)
            self.dialogs.append(dialog)
            dialog.show()
    
    def submitContact(self):
        self.close()

class Window(QDialog):
    def __init__(self, parent=None):
        super(Window, self).__init__(parent)
        self.createFormGroupBox()
        
    def createFormGroupBox(self):
        # a figure instance to plot on
        self.figure = Figure()
        self.setMinimumSize(QSize(940, 450))
        
        font = QFont()
        font.setPointSize(12)
        
        hboxa = QVBoxLayout()
        hboxa1 = QVBoxLayout()
        hboxa2 = QVBoxLayout()
        hboxa3 = QVBoxLayout()
        
        buttons = QHBoxLayout()
        
        aixtron = QLabel("Aixtron")
        amat = QLabel("Applied Materials")
        asm = QLabel("ASM Internaitonal")
        eugene = QLabel("Eugene Technology")
        hitachi = QLabel("Hitachi Kokusai Electric")
        jusung = QLabel("Jusung Engineering")
        lam = QLabel("Lam Research")
        orbotech = QLabel("Orbotech")
        tes = QLabel("Tes")
        tokyo = QLabel("Tokyo Electron")
        ultratech = QLabel("Ultratech")
        veeco = QLabel("Veeco")
        wonik = QLabel("Wonik IPS")
        select_all_comp = QLabel("Select all:")
        
        nlpcvd = QLabel("NLPCVD")
        pcvd = QLabel("PCVD")
        tcvd = QLabel("TCVD")
        aldp = QLabel("ALDP")
                     
        # size policy
        not_resize = amat.sizePolicy();
        not_resize.setRetainSizeWhenHidden(True);
        amat.setSizePolicy(not_resize);
        amat.setVisible(True)
                
        self.setFont(font)
        
        self.select_all_comp = QCheckBox()
        self.select_all_comp.setFixedSize(20,26)
        self.aixtron = QCheckBox()
        self.aixtron.setFixedSize(20,25)
        self.amat = QCheckBox()
        self.amat.setFixedSize(20,25)
        self.asm = QCheckBox()
        self.asm.setFixedSize(20,25)
        self.eugene = QCheckBox()
        self.eugene.setFixedSize(20,25)
        self.hitachi = QCheckBox()
        self.hitachi.setFixedSize(20,25)
        self.lam = QCheckBox()
        self.lam.setFixedSize(20,25)
        self.jusung = QCheckBox()
        self.jusung.setFixedSize(20,25)
        self.orbotech = QCheckBox()
        self.orbotech.setFixedSize(20,25)
        self.tes = QCheckBox()
        self.tes.setFixedSize(20,25)
        self.tokyo = QCheckBox()
        self.tokyo.setFixedSize(20,25)
        self.ultratech = QCheckBox()
        self.ultratech.setFixedSize(20,25)
        self.veeco = QCheckBox()
        self.veeco.setFixedSize(20,25)
        self.wonik = QCheckBox()
        self.wonik.setFixedSize(20,25)
        
        self.nlpcvd = QCheckBox()
        self.pcvd = QCheckBox()
        self.tcvd = QCheckBox()
        self.aldp = QCheckBox()

        self.ratio_box = QCheckBox()
        
        self.finish_graph = QPushButton("Generate graph!")
        #self.finish_graph.clicked.connect(self.generate_graph)
        
        self.select_all_comp.stateChanged.connect(self.onStateChange)
        self.select_all_comp.setChecked(True)
        
        self.amat.stateChanged.connect(self.onStateChange)
        self.amat.setChecked(False)
        
        self.aixtron.stateChanged.connect(self.onStateChange)
        self.aixtron.setChecked(False)
        
        self.asm.stateChanged.connect(self.onStateChange)
        self.asm.setChecked(False)
        
        self.eugene.stateChanged.connect(self.onStateChange)
        self.eugene.setChecked(False)
        
        self.hitachi.stateChanged.connect(self.onStateChange)
        self.hitachi.setChecked(False)
        
        self.jusung.stateChanged.connect(self.onStateChange)
        self.jusung.setChecked(False)
        
        self.orbotech.stateChanged.connect(self.onStateChange)
        self.orbotech.setChecked(False)
        
        self.ultratech.stateChanged.connect(self.onStateChange)
        self.ultratech.setChecked(False)
        
        self.veeco.stateChanged.connect(self.onStateChange)
        self.veeco.setChecked(False)
        
        self.wonik.stateChanged.connect(self.onStateChange)
        self.wonik.setChecked(False)
        
        self.lam.stateChanged.connect(self.onStateChange)
        self.lam.setChecked(False)
        
        self.tes.stateChanged.connect(self.onStateChange)
        self.tes.setChecked(False)
        
        self.tokyo.stateChanged.connect(self.onStateChange)
        self.tokyo.setChecked(False)
        
        self.nlpcvd.stateChanged.connect(self.onStateChange)
        self.nlpcvd.setChecked(False)
        
        self.pcvd.stateChanged.connect(self.onStateChange)
        self.pcvd.setChecked(False)
        
        self.tcvd.stateChanged.connect(self.onStateChange)
        self.tcvd.setChecked(False)
        
        self.aldp.stateChanged.connect(self.onStateChange)
        self.aldp.setChecked(False)
        
        # this is the Canvas Widget that displays the `figure`
        # it takes the `figure` instance as a parameter to __init__
        self.canvas = FigureCanvas(self.figure)

        # this is the Navigation widget
        # it takes the Canvas widget and a parent
        self.toolbar = NavigationToolbar(self.canvas, self)

        # Just some button connected to `plot` method
        self.button = QPushButton('Generate Graph')
        self.button.clicked.connect(self.plot)

        # set the layout

        layout = QHBoxLayout()

        layout2 = QVBoxLayout()
        layout2.addWidget(self.toolbar)
        
        hboxa.addWidget(select_all_comp)
        hboxa.addWidget(aixtron)
        hboxa.addWidget(amat)
        hboxa.addWidget(asm)
        hboxa.addWidget(eugene)
        hboxa.addWidget(hitachi)
        hboxa.addWidget(jusung)
        hboxa.addWidget(lam)
        hboxa.addWidget(orbotech)
        hboxa.addWidget(tes)
        hboxa.addWidget(tokyo)
        hboxa.addWidget(ultratech)
        hboxa.addWidget(veeco)
        hboxa.addWidget(wonik)
        layout.addLayout(hboxa)
        
        hboxa2.addWidget(self.select_all_comp)
        hboxa2.addWidget(self.aixtron)
        hboxa2.addWidget(self.amat)
        hboxa2.addWidget(self.asm)
        hboxa2.addWidget(self.eugene)
        hboxa2.addWidget(self.hitachi)
        hboxa2.addWidget(self.jusung)
        hboxa2.addWidget(self.lam)
        hboxa2.addWidget(self.orbotech)
        hboxa2.addWidget(self.tes)
        hboxa2.addWidget(self.tokyo)
        hboxa2.addWidget(self.ultratech)
        hboxa2.addWidget(self.veeco)
        hboxa2.addWidget(self.wonik)
        layout.addLayout(hboxa2)
        
        
        layout2.addWidget(self.canvas)
        layout.addLayout(layout2)
        
        hboxa1.addWidget(nlpcvd)
        hboxa1.addWidget(self.nlpcvd)
        hboxa1.addWidget(pcvd)
        hboxa1.addWidget(self.pcvd)
        hboxa1.addWidget(aldp)
        hboxa1.addWidget(self.aldp)
        hboxa1.addWidget(tcvd)
        hboxa1.addWidget(self.tcvd)
        layout.addLayout(hboxa1)
        
        layout2.addWidget(self.button)
        
        self.setLayout(layout)

    def onStateChange(self, state):
        if state == Qt.Unchecked:
            if self.sender() == self.select_all_comp:
                self.amat.setChecked(True)
                self.aixtron.setChecked(True)
                self.asm.setChecked(True)
                self.eugene.setChecked(True)
                self.hitachi.setChecked(True)
                self.lam.setChecked(True)
                self.veeco.setChecked(True)
                self.wonik.setChecked(True)
                self.tes.setChecked(True)
                self.orbotech.setChecked(True)
                self.ultratech.setChecked(True)
                self.jusung.setChecked(True)
                self.tokyo.setChecked(True)
                self.nlpcvd.setChecked(True)
                self.aldp.setChecked(True)
                self.pcvd.setChecked(True)
                self.tcvd.setChecked(True)

        elif state == Qt.Checked:
            if self.sender() == self.select_all_comp:
                self.amat.setChecked(False)
                self.aixtron.setChecked(False)
                self.asm.setChecked(False)
                self.eugene.setChecked(False)
                self.hitachi.setChecked(False)
                self.lam.setChecked(False)
                self.veeco.setChecked(False)
                self.wonik.setChecked(False)
                self.tes.setChecked(False)
                self.ultratech.setChecked(False)
                self.jusung.setChecked(False)
                self.tokyo.setChecked(False)
                self.orbotech.setChecked(False)
                self.nlpcvd.setChecked(False)
                self.aldp.setChecked(False)
                self.pcvd.setChecked(False)
                self.tcvd.setChecked(False)

    def plot(self):
        
        ''' plot some random stuff '''
        # random data

        # create an axis
        ax = self.figure.add_subplot(111)
        
        # discards the old graph
        ax.clear()

        # plot data
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2']]
        
        if self.nlpcvd.isChecked() and self.pcvd.isChecked() and self.aldp.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 =691.3738
                amat_total_14 = 1057.973
                amat_total_15 = 1182.857
                amat_total_16 = 1782.115
                amat_total_17 = 2426.448
                
                amat_eq = ['Atomic Layer Deposition Platforms', 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 0)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            else:
                amat_total = ["","","","","",""]
                
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 580.4183
                lam_total_14 = 805.3354
                lam_total_15 = 901.35
                lam_total_16 = 1151.953
                lam_total_17 = 2184.25
                
                lam_eq = ['Atomic Layer Deposition Platforms', 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            else:
                lam_total = ["","","","","",""]
                
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 107.9
                wonik_total_14 = 124.0
                wonik_total_15 = 128.4769
                wonik_total_16 = 137.2021
                wonik_total_17 = 372.1
                
                wonik_eq = ['Atomic Layer Deposition Platforms', 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
            
            else:
                wonik_total = ["","","","","",""]
            
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total,'Wonik':wonik_total}, index = year)
            
            df[['Amat', 'Lam','Wonik']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-50)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
        
        elif self.nlpcvd.isChecked() and self.pcvd.isChecked() and self.tcvd.isChecked():
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 406.9015
                tokyo_total_14 = 643.9228
                tokyo_total_15 = 663.3027
                tokyo_total_16 = 583.8111
                tokyo_total_17 = 1014.6
                
                tokyo_eq = ['Nontube LPCVD', 'Plasma CVD','Tube CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
        
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Tokyo':tokyo_total}, index = year)
            
            df[['Tokyo']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-50)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
        
        elif self.tcvd.isChecked() and self.aldp.isChecked() and self.pcvd.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 401.5537
                asm_total_14 = 500.5012
                asm_total_15 = 520.3123
                asm_total_16 = 429.0665
                asm_total_17 = 526.7842
                
                asm_eq = ['Atomic Layer Deposition Platforms','Plasma CVD','Tube CVD' ]
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
        
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Asm':asm_total}, index = year)
            
            df[['Asm']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-50)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
                
        elif self.nlpcvd.isChecked() and self.pcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 =691.3738
                amat_total_14 = 1057.973
                amat_total_15 = 1156.762
                amat_total_16 = 1686.915
                amat_total_17 = 2365.264
                
                amat_eq = ['Nontube LPCVD', 'Plasma CVD']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 1)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            else:
                amat_total = ["","","","","",""]
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 58.0
                eugene_total_14 = 35.9
                eugene_total_15 = 42.1
                eugene_total_16 = 38.779
                eugene_total_17 = 48.200
                
                eugene_eq = ['Nontube LPCVD', 'Plasma CVD']
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = (i['Yearly'])
                    eugene.append(ii)
                    
                eugene_total_18 = round(sum(eugene), 1)
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
            else:
                eugene_total= ["","","","","",""]
            
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 580.4183
                lam_total_14 = 759.7654
                lam_total_15 = 805.95
                lam_total_16 = 1088.365
                lam_total_17 = 1969.35
                
                lam_eq = [ 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),1)
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]    
            else:
                lam_total = ["","","","","",""]
            if self.tes.isChecked():
                tes = []
                tes_total_13 = 28.0
                tes_total_14 = 72.6
                tes_total_15 = 68.4
                tes_total_16 = 115.301
                tes_total_17 = 180.6
                
                tes_eq = [ 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in tes_eq:
                    i = df[(df.Competitor == 'Tes' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tes.append(ii)
                
                tes_total_18 = round(sum(tes),1)
                
                tes_total = [tes_total_13,tes_total_14, tes_total_15, tes_total_16, tes_total_17, tes_total_18]
            else:
                tes_total = ["","","","","",""]
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 104.4687
                tokyo_total_14 = 204.9875
                tokyo_total_15 = 246.966
                tokyo_total_16 = 237.7651
                tokyo_total_17 = 439.3
                
                tokyo_eq = ['Nontube LPCVD', 'Plasma CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),1)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
            else:
                tokyo_total = ["","","","","",""]
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 68.5
                wonik_total_14 = 78.7
                wonik_total_15 = 80.7259
                wonik_total_16 = 91.5999
                wonik_total_17 = 248.4
                
                wonik_eq = [ 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),1)
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
            else:
                wonik_total = ["","","","","",""]
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total,'Wonik':wonik_total, 'Eugene':eugene_total, 'Tes':tes_total, 'Tokyo':tokyo_total}, index = year)
            
            df[['Amat', 'Lam','Wonik', 'Eugene', 'Tes', 'Tokyo']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-50)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
            
        elif self.nlpcvd.isChecked() and self.aldp.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 =72.141341
                amat_total_14 = 133.818248
                amat_total_15 = 114.9833
                amat_total_16 = 343.8527
                amat_total_17 = 209.4539
                
                amat_eq = ['Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 0)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            else:
                amat_total = ["","","","","",""]
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 125.6684703
                lam_total_14 = 224.2045
                lam_total_15 = 303.75
                lam_total_16 = 283.3968
                lam_total_17 = 735.1
                
                lam_eq = [ 'Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            else:
                lam_total = ["","","","","",""]
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 39.4
                wonik_total_14 = 45.3
                wonik_total_15 = 47.75091049
                wonik_total_16 = 75.60211945
                wonik_total_17 = 174.9
                
                wonik_eq = [ 'Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
            else:
                wonik_total = ["","","","","",""]
                
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 44.1
                eugene_total_14 = 27.3
                eugene_total_15 = 29.9
                eugene_total_16 = 38.77942
                eugene_total_17 = 60.5
                
                eugene_eq = [ 'Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_total = [eugene_total_13,eugene_total_14, eugene_total_15,eugene_total_16, eugene_total_17, eugene_total_18]
            else:
                eugene_total = ["","","","","",""]
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total,'Eugene':eugene_total,'Wonik':wonik_total}, index = year)
            
            df[['Amat', 'Lam','Eugene','Wonik']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-10)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
                
        elif self.tcvd.isChecked() and self.pcvd.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 53.5832
                asm_total_14 = 54.3012
                asm_total_15 = 51.8123
                asm_total_16 = 39.386
                asm_total_17 = 80.6
                
                asm_eq = ['Tube CVD', 'Plasma CVD']
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                    
                asm_total_18 = round(sum(asm), 1)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16,asm_total_17, asm_total_18]
            else:
                asm_total = ["","","","","",""]
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 302.432679
                tokyo_total_14 = 438.935
                tokyo_total_15 = 416.3365
                tokyo_total_16 = 350.2369
                tokyo_total_17 = 583.5
                
                tokyo_eq = ['Tube CVD', 'Plasma CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
            
            else:
                tokyo_total = ["","","","","",""]
                
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'ASM':asm_total, 'Tokyo':tokyo_total}, index = year)
            
            df[['ASM', 'Tokyo']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-20)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
        
        elif self.tcvd.isChecked() and self.aldp.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 371.2758
                asm_total_14 = 469.79
                asm_total_15 = 495.8
                asm_total_16 = 411.959
                asm_total_17 = 473.5842
                
                asm_eq = ['Tube CVD', 'Atomic Layer Deposition Platforms']
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                    
                asm_total_18 = round(sum(asm), 1)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16,asm_total_17, asm_total_18]
            else:
                asm_total = ["","","","","",""]
            
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
            df = pd.DataFrame({'ASM':asm_total}, index = year)
            df[['ASM']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-50)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
                
        elif self.nlpcvd.isChecked() and self.tcvd.isChecked():
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 406.9015
                tokyo_total_14 = 643.9228
                tokyo_total_15 = 663.3027
                tokyo_total_16 = 579.6202
                tokyo_total_17 = 1006.4
                
                tokyo_eq = ['Tube CVD', 'Nontube LPCVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
            
            else:
                tokyo_total =["","","","","",""]
                
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Tokyo':tokyo_total}, index = year)
            
            df[['Tokyo']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-50)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
        elif self.aldp.isChecked() and self.pcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 619.232479
                amat_total_14 = 924.1551212
                amat_total_15 = 1093.968
                amat_total_16 = 1533.462
                amat_total_17 = 2278.178
                
                amat_eq = ['Atomic Layer Deposition Platforms', 'Plasma CVD']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 1)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            else:
                amat_total = ["","","","","",""]
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 378.2477
                asm_total_14 = 476.9612
                asm_total_15 = 493.0123
                asm_total_16 = 406.788
                asm_total_17 = 499.3842
                
                asm_eq = ['Atomic Layer Deposition Platforms', 'Plasma CVD']
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                    
                asm_total_18 = round(sum(asm), 1)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16,asm_total_17, asm_total_18]
            else:
                asm_total = ["","","","","",""]
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 580.4183
                lam_total_14 = 759.7654
                lam_total_15 = 805.95
                lam_total_16 = 1088.365
                lam_total_17 = 1969.35
                
                lam_eq = [ 'Atomic Layer Deposition Platforms', 'Plasma CVD']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),1)
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            else:
                lam_total = ["","","","","",""]
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 107.9
                wonik_total_14 = 124.0
                wonik_total_15 = 128.4769
                wonik_total_16 = 107.2021
                wonik_total_17 = 320.7
                
                wonik_eq = [ 'Atomic Layer Deposition Platforms', 'Plasma CVD']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),1)
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
            else:
                wonik_total = ["","","","","",""]
                
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 13.9
                eugene_total_14 = 8.6
                eugene_total_15 = 12.2
                eugene_total_16 = 0
                eugene_total_17 = 12.300
                
                eugene_eq = [ 'Plasma CVD', 'Atomic Layer Deposition Platforms']
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_total = [eugene_total_13,eugene_total_14, eugene_total_15,eugene_total_16, eugene_total_17, eugene_total_18]
            else:
                eugene_total = ["","","","","",""]
        
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
    
            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total,'Wonik':wonik_total, 'Eugene':eugene_total,'ASM':asm_total}, index = year)
            
            df[['Amat', 'Lam','Eugene','Wonik', 'ASM']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-7)))
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")
        
        elif self.nlpcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 72.1413
                amat_total_14 = 133.818
                amat_total_15 = 88.889
                amat_total_16 = 248.652
                amat_total_17 = 148.270
                amat_eq = ['Nontube LPCVD' ]
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                
                amat_total_18 = round(sum(amat),0)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
                
            else:
                amat_total=["","","","","",""]
                
            if self.eugene.isChecked():
                
                eugene = []
                eugene_total_13 = 44.1
                eugene_total_14 = 27.3
                eugene_total_15 = 29.9
                eugene_total_16 = 38.7794
                eugene_total_17 = 48.200
                
                eugene_eq = ['Nontube LPCVD' ]
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                    
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
            
            
            else:
                eugene_total= ["","","","","",""]
                
                
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 125.6684
                lam_total_14 = 178.6345
                lam_total_15 = 208.35
                lam_total_16 = 219.8092499
                lam_total_17 = 520.0999
                
                lam_eq = ['Nontube LPCVD' ]
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_total = [lam_total_13, lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            
            else:
                lam_total = ["","","","","",""]
            
            if self.tes.isChecked():
                tes = []
                tes_total_13 = 0.0
                tes_total_14 = 0.0
                tes_total_15 = 0.0
                tes_total_16 = 40.001
                tes_total_17 = 76.400
                
                tes_eq = ['Nontube LPCVD' ]
                
                for eq in tes_eq:
                    i = df[(df.Competitor == 'Tes' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tes.append(ii)
                
                tes_total_18 = round(sum(tes),0)
                
                tes_total = [tes_total_13, tes_total_14, tes_total_15, tes_total_16, tes_total_17, tes_total_18]
            
            else:
                tes_total = ["","","","","",""]
                
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 104.4687
                tokyo_total_14 = 204.98759
                tokyo_total_15 = 246.966132
                tokyo_total_16 = 233.5742035
                tokyo_total_17 = 431.1000
                
                tokyo_eq = ['Nontube LPCVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
            
            else:
                tokyo_total = ["","","","","",""]
            
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 0.0
                wonik_total_14 = 0.0
                wonik_total_15 = 0.0
                wonik_total_16 = 30.0
                wonik_total_17 = 51.2999
                
                wonik_eq = ['Nontube LPCVD' ]
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_total = [wonik_total_13, wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]
            else:
                wonik_total = ["","","","","",""]
                
            if self.aixtron.isChecked():
                aixtron = []
                aixtron_total_13 = 0.0
                aixtron_total_14 = 0.0
                aixtron_total_15 = 0.0
                aixtron_total_16 = 0.0
                aixtron_total_17 = 0.8999
                
                aixtron_eq = ['Nontube LPCVD' ]
                
                for eq in aixtron_eq:
                    i = df[(df.Competitor == 'Aixtron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    aixtron.append(ii)
                
                aixtron_total_18 = round(sum(aixtron),1)
                
                aixtron_total = [aixtron_total_13, aixtron_total_14, aixtron_total_15, aixtron_total_16, aixtron_total_17, aixtron_total_18]
            
            else:
                aixtron_total = ["","","","","",""]
                
            if self.hitachi.isChecked():
                hitachi_total_13 = 0.0
                hitachi_total_14 = 0.0
                hitachi_total_15 = 2.925022
                hitachi_total_16 = 0.0
                hitachi_total_17 = 0.0
                hitachi_total_18 = 0.0
                
                
                hitachi_total = [hitachi_total_13, hitachi_total_14, hitachi_total_15, hitachi_total_16, hitachi_total_17, hitachi_total_18]        
    
            else: 
                hitachi_total= ["","","","","",""]
    
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']

            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total, 'Hitachi':hitachi_total,'Aixtron':aixtron_total,'Wonik':wonik_total,'Eugene':eugene_total,'Tes':tes_total, 'Tokyo':tokyo_total}, index = year)
            
            df[['Amat', 'Lam','Wonik', 'Eugene', 'Tes', 'Tokyo', 'Aixtron', 'Hitachi']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-5)), fontsize = 5)
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")     
            
        elif self.pcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 619.232479
                amat_total_14 = 924.155121
                amat_total_15 = 1067.973379
                amat_total_16 = 1438.2618
                amat_total_17 = 2216.994
                
                amat_eq = ['Plasma CVD' ]
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                
                amat_total_18 = round(sum(amat),0)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            else:
                amat_total = ["","","","","",""]
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 30.2778
                asm_total_14 = 30.7612
                asm_total_15 = 24.5123
                asm_total_16 = 17.1075
                asm_total_17 = 53.2000
                
                asm_eq = ['Plasma CVD' ]
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
            else:
                asm_total = ["","","","","",""]
            if self.eugene.isChecked():
                eugene_total_13 = 13.9
                eugene_total_14 = 8.6
                eugene_total_15 = 12.2
                eugene_total_16 = 0.0
                eugene_total_17 = 0.0
                eugene_total_18 = 0.0
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]

            else:
                eugene_total = ["","","","","",""]
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 454.749785
                lam_total_14 = 581.1309032
                lam_total_15 = 597.6
                lam_total_16 = 868.5560303
                lam_total_17 = 1449.25
                
                lam_eq = ['Plasma CVD' ]
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_total = [lam_total_13, lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            else:
                lam_total = ["","","","","",""]
            if self.orbotech.isChecked():
                orbotech = []
                orbotech_total_13 = 0.0
                orbotech_total_14 = 0.0
                orbotech_total_15 = 15.5496
                orbotech_total_16 = 17.0734
                orbotech_total_17 = 23.7000
                
                orbotech_eq = ['Plasma CVD' ]
                
                for eq in orbotech_eq:
                    i = df[(df.Competitor == 'Orbotech' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    orbotech.append(ii)
                
                orbotech_total_18 = round(sum(orbotech),0)
                
                orbotech_total = [orbotech_total_13, orbotech_total_14, orbotech_total_15, orbotech_total_16, orbotech_total_17, orbotech_total_18]
            else:
                orbotech_total = ["","","","","",""]
            if self.tes.isChecked():
                tes = []
                tes_total_13 = 28.0
                tes_total_14 = 72.6
                tes_total_15 = 68.4
                tes_total_16 = 75.3000
                tes_total_17 = 104.19999
                
                tes_eq = ['Plasma CVD' ]
                
                for eq in tes_eq:
                    i = df[(df.Competitor == 'Tes' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tes.append(ii)
                
                tes_total_18 = round(sum(tes),0)
                
                tes_total = [tes_total_13, tes_total_14, tes_total_15, tes_total_16, tes_total_17, tes_total_18]
            else:
                tes_total = ["","","","","",""]
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 0.0
                tokyo_total_14 = 0.0
                tokyo_total_15 = 0.0
                tokyo_total_16 = 4.0
                tokyo_total_17 = 8.0
                
                tokyo_eq = ['Plasma CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
            else:
                tokyo_total = ["","","","","",""]
            
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 68.5
                wonik_total_14 = 78.7
                wonik_total_15 = 80.7259
                wonik_total_16 = 61.5999
                wonik_total_17 = 197.100
                
                wonik_eq = ['Plasma CVD' ]
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik .append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
            
                wonik_total = [wonik_total_13, wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]
            else:
                wonik_total = ["","","","","",""]
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']

            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total, 'Asm':asm_total,'Orbotech':orbotech_total,'Wonik':wonik_total,'Eugene':eugene_total,'Tes':tes_total, 'Tokyo':tokyo_total}, index = year)
            
            df[['Amat', 'Lam','Wonik', 'Eugene', 'Tes', 'Tokyo', 'Orbotech', 'Asm']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-10)), fontsize = 6)
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")  
                
                
        elif self.aldp.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 0.0
                amat_total_14 = 0.0
                amat_total_15 = 26.09425
                amat_total_16 = 95.19999
                amat_total_17 = 61.183925
                
                amat_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                
                amat_total_18 = round(sum(amat),0)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            else:
                amat_total = ["","","","","",""]
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 0.0
                lam_total_14 = 45.5700295
                lam_total_15 = 95.4
                lam_total_16 = 63.5875
                lam_total_17 = 214.8999
                
                lam_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            else:
                lam_total = ["","","","","",""]
        
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 39.4
                wonik_total_14 = 45.3
                wonik_total_15 = 47.7509
                wonik_total_16 = 45.60211
                wonik_total_17 = 123.59999
                
                wonik_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_total = [wonik_total_13, wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]
            else:
                wonik_total = ["","","","","",""]
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 0
                eugene_total_14 = 0
                eugene_total_15 = 0
                eugene_total_16 = 0
                eugene_total_17 = 12.30000
                
                eugene_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
            else:
                eugene_total = ["","","","","",""]
            if self.veeco.isChecked():
                veeco = []
                veeco_total_13 = 0.0
                veeco_total_14 = 0.0
                veeco_total_15 = 0.0
                veeco_total_16 = 0.0
                veeco_total_17 = 4.115830898
                
                veeco_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in veeco_eq:
                    i = df[(df.Competitor == 'Veeco' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    veeco.append(ii)
                
                veeco_total_18 = round(sum(veeco),0)
                
                veeco_total = [veeco_total_13, veeco_total_14, veeco_total_15, veeco_total_16, veeco_total_17, veeco_total_18]
            else:
                veeco_total = ["","","","","",""]
            if self.ultratech.isChecked():
                ultra = []
                ultra_total_13 = 7.225
                ultra_total_14 = 5.155
                ultra_total_15 = 8.75
                ultra_total_16 = 9.1300
                ultra_total_17 = 7.4555
                
                ultra_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in ultra_eq:
                    i = df[(df.Competitor == 'Ultratech' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    ultra.append(ii)
                
                ultra_total_18 = round(sum(ultra),0)

                ultra_total = [ultra_total_13, ultra_total_14, ultra_total_15, ultra_total_16, ultra_total_17, ultra_total_18]
            else:
                ultra_total = ["","","","","",""]
            if self.aixtron.isChecked():
                aixtron = []
                aixtron_total_13 = 41.2336
                aixtron_total_14 = 20.856
                aixtron_total_15 = 33.0
                aixtron_total_16 = 27.059999
                aixtron_total_17 = 39.7999
                
                aixtron_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in aixtron_eq:
                    i = df[(df.Competitor == 'Aixtron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    aixtron.append(ii)
                
                aixtron_total_18 = round(sum(aixtron),0)
                
                aixtron_total = [aixtron_total_13, aixtron_total_14, aixtron_total_15, aixtron_total_16, aixtron_total_17, aixtron_total_18]
            else:
                aixtron_total =["","","","","",""]
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 347.9698445
                asm_total_14 = 446.2
                asm_total_15 = 468.5
                asm_total_16 = 389.6805115
                asm_total_17 = 446.1841736
                
                asm_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
            
            else:
                asm_total = ["","","","","",""]
            if self.jusung.isChecked():
                jusung = []
                jusung_total_13 = 32.0
                jusung_total_14 = 51.0
                jusung_total_15 = 63.1
                jusung_total_16 = 86.900
                jusung_total_17 = 97.80
                
                jusung_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in jusung_eq:
                    i = df[(df.Competitor == 'Jusung Engineering' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    jusung.append(ii)
                
                jusung_total_18 = round(sum(jusung),0)
                
                jusung_total = [jusung_total_13, jusung_total_14, jusung_total_15, jusung_total_16, jusung_total_17, jusung_total_18]
        
            else:
                jusung_total = ["","","","","",""]
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']

            df = pd.DataFrame({'Amat':amat_total, 'Lam':lam_total,'Wonik':wonik_total,'Eugene':eugene_total,'Veeco':veeco_total, 'Ultratech':ultra_total, 'Aixtron':aixtron_total, 'ASM':asm_total, 'Jusung':jusung_total}, index = year)
            
            df[['Amat', 'Lam','Wonik', 'Eugene', 'Veeco', 'Ultratech', 'Aixtron', 'ASM', 'Jusung']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-10)), fontsize = 6)
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")  
                
        elif self.tcvd.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 23.30597
                asm_total_14 = 23.54
                asm_total_15 = 27.3
                asm_total_16 = 22.278499
                asm_total_17 = 27.399999
                
                asm_eq = ['Tube CVD' ]
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
            else:
                asm_total = ["","","","","",""]
            if self.hitachi.isChecked():
                hitachi = []
                hitachi_total_13 = 318.3580273
                hitachi_total_14 = 479.1185744
                hitachi_total_15 = 504.640823
                hitachi_total_16 = 412.2118835
                hitachi_total_17 = 776.1200562
                
                hitachi_eq = ['Tube CVD' ]
                
                for eq in hitachi_eq:
                    i = df[(df.Competitor == 'Hitachi Kokusai Electric' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    hitachi.append(ii)
                
                hitachi_total_18 = round(sum(hitachi),0)
                
                hitachi_total = [hitachi_total_13, hitachi_total_14, hitachi_total_15, hitachi_total_16, hitachi_total_17, hitachi_total_18]
            else:
                hitachi_total = ["","","","","",""]
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 302.4326797
                tokyo_total_14 = 438.9352201
                tokyo_total_15 = 416.3365286
                tokyo_total_16 = 346.0460205
                tokyo_total_17 = 575.2999878
                
                tokyo_eq = ['Tube CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
            else:
                tokyo_total = ["","","","","",""]
            year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']

            df = pd.DataFrame({ 'ASM':asm_total, 'Hitachi':hitachi_total, 'Tokyo':tokyo_total}, index = year)
            
            df[['ASM', 'Hitachi', 'Tokyo']].plot.bar(stacked=True, colormap="rainbow", ax=ax, alpha=0.7)
            for p in ax.patches:
                width, height = p.get_width(), p.get_height()
                x, y = p.get_xy() 
                ax.annotate('{:.0f} '.format(height), (p.get_x()+.15*width, p.get_y()+.4*(height-10)), fontsize = 6)
                ax.set_title("Segment competitor by equipment")
                ax.set_xlabel("Year")
                ax.set_ylabel("Revenue (in millions)")  
        # refresh canvas
        self.canvas.draw()
    
    
class line(QDialog):
    def __init__(self, parent=None):
        super(line, self).__init__(parent)
        self.createFormGroupBox()
        
    def createFormGroupBox(self):
        # a figure instance to plot on
        self.figure = Figure()
        self.setMinimumSize(QSize(940, 450))
        
        font = QFont()
        font.setPointSize(12)
        
        hboxa = QVBoxLayout()
        hboxa1 = QVBoxLayout()
        hboxa2 = QVBoxLayout()
        hboxa3 = QVBoxLayout()
        
        buttons = QHBoxLayout()
        
        aixtron = QLabel("Aixtron")
        amat = QLabel("Applied Materials")
        asm = QLabel("ASM Internaitonal")
        eugene = QLabel("Eugene Technology")
        hitachi = QLabel("Hitachi Kokusai Electric")
        jusung = QLabel("Jusung Engineering")
        lam = QLabel("Lam Research")
        orbotech = QLabel("Orbotech")
        tes = QLabel("Tes")
        tokyo = QLabel("Tokyo Electron")
        ultratech = QLabel("Ultratech")
        veeco = QLabel("Veeco")
        wonik = QLabel("Wonik IPS")
        select_all_comp = QLabel("Select all:")
        
        nlpcvd = QLabel("NLPCVD")
        pcvd = QLabel("PCVD")
        tcvd = QLabel("TCVD")
        aldp = QLabel("ALDP")
                     
        # size policy
        not_resize = amat.sizePolicy();
        not_resize.setRetainSizeWhenHidden(True);
        amat.setSizePolicy(not_resize);
        amat.setVisible(True)
                
        self.setFont(font)
        
        self.select_all_comp = QCheckBox()
        self.select_all_comp.setFixedSize(20,26)
        self.aixtron = QCheckBox()
        self.aixtron.setFixedSize(20,25)
        self.amat = QCheckBox()
        self.amat.setFixedSize(20,25)
        self.asm = QCheckBox()
        self.asm.setFixedSize(20,25)
        self.eugene = QCheckBox()
        self.eugene.setFixedSize(20,25)
        self.hitachi = QCheckBox()
        self.hitachi.setFixedSize(20,25)
        self.lam = QCheckBox()
        self.lam.setFixedSize(20,25)
        self.jusung = QCheckBox()
        self.jusung.setFixedSize(20,25)
        self.orbotech = QCheckBox()
        self.orbotech.setFixedSize(20,25)
        self.tes = QCheckBox()
        self.tes.setFixedSize(20,25)
        self.tokyo = QCheckBox()
        self.tokyo.setFixedSize(20,25)
        self.ultratech = QCheckBox()
        self.ultratech.setFixedSize(20,25)
        self.veeco = QCheckBox()
        self.veeco.setFixedSize(20,25)
        self.wonik = QCheckBox()
        self.wonik.setFixedSize(20,25)
        
        self.nlpcvd = QCheckBox()
        self.pcvd = QCheckBox()
        self.tcvd = QCheckBox()
        self.aldp = QCheckBox()

        self.ratio_box = QCheckBox()
        
        self.finish_graph = QPushButton("Generate graph!")
        #self.finish_graph.clicked.connect(self.generate_graph)
        
        self.select_all_comp.stateChanged.connect(self.onStateChange)
        self.select_all_comp.setChecked(True)
        
        self.amat.stateChanged.connect(self.onStateChange)
        self.amat.setChecked(False)
        
        self.aixtron.stateChanged.connect(self.onStateChange)
        self.aixtron.setChecked(False)
        
        self.asm.stateChanged.connect(self.onStateChange)
        self.asm.setChecked(False)
        
        self.eugene.stateChanged.connect(self.onStateChange)
        self.eugene.setChecked(False)
        
        self.hitachi.stateChanged.connect(self.onStateChange)
        self.hitachi.setChecked(False)
        
        self.jusung.stateChanged.connect(self.onStateChange)
        self.jusung.setChecked(False)
        
        self.orbotech.stateChanged.connect(self.onStateChange)
        self.orbotech.setChecked(False)
        
        self.ultratech.stateChanged.connect(self.onStateChange)
        self.ultratech.setChecked(False)
        
        self.veeco.stateChanged.connect(self.onStateChange)
        self.veeco.setChecked(False)
        
        self.wonik.stateChanged.connect(self.onStateChange)
        self.wonik.setChecked(False)
        
        self.lam.stateChanged.connect(self.onStateChange)
        self.lam.setChecked(False)
        
        self.tes.stateChanged.connect(self.onStateChange)
        self.tes.setChecked(False)
        
        self.tokyo.stateChanged.connect(self.onStateChange)
        self.tokyo.setChecked(False)
        
        self.nlpcvd.stateChanged.connect(self.onStateChange)
        self.nlpcvd.setChecked(False)
        
        self.pcvd.stateChanged.connect(self.onStateChange)
        self.pcvd.setChecked(False)
        
        self.tcvd.stateChanged.connect(self.onStateChange)
        self.tcvd.setChecked(False)
        
        self.aldp.stateChanged.connect(self.onStateChange)
        self.aldp.setChecked(False)
        
        # this is the Canvas Widget that displays the `figure`
        # it takes the `figure` instance as a parameter to __init__
        self.canvas = FigureCanvas(self.figure)

        # this is the Navigation widget
        # it takes the Canvas widget and a parent
        self.toolbar = NavigationToolbar(self.canvas, self)

        # Just some button connected to `plot` method
        self.button = QPushButton('Generate Graph')
        self.button.clicked.connect(self.plot)

        # set the layout

        layout = QHBoxLayout()

        layout2 = QVBoxLayout()
        layout2.addWidget(self.toolbar)
        
        hboxa.addWidget(select_all_comp)
        hboxa.addWidget(aixtron)
        hboxa.addWidget(amat)
        hboxa.addWidget(asm)
        hboxa.addWidget(eugene)
        hboxa.addWidget(hitachi)
        hboxa.addWidget(jusung)
        hboxa.addWidget(lam)
        hboxa.addWidget(orbotech)
        hboxa.addWidget(tes)
        hboxa.addWidget(tokyo)
        hboxa.addWidget(ultratech)
        hboxa.addWidget(veeco)
        hboxa.addWidget(wonik)
        layout.addLayout(hboxa)
        
        hboxa2.addWidget(self.select_all_comp)
        hboxa2.addWidget(self.aixtron)
        hboxa2.addWidget(self.amat)
        hboxa2.addWidget(self.asm)
        hboxa2.addWidget(self.eugene)
        hboxa2.addWidget(self.hitachi)
        hboxa2.addWidget(self.jusung)
        hboxa2.addWidget(self.lam)
        hboxa2.addWidget(self.orbotech)
        hboxa2.addWidget(self.tes)
        hboxa2.addWidget(self.tokyo)
        hboxa2.addWidget(self.ultratech)
        hboxa2.addWidget(self.veeco)
        hboxa2.addWidget(self.wonik)
        layout.addLayout(hboxa2)
        
        
        layout2.addWidget(self.canvas)
        layout.addLayout(layout2)
        
        hboxa1.addWidget(nlpcvd)
        hboxa1.addWidget(self.nlpcvd)
        hboxa1.addWidget(pcvd)
        hboxa1.addWidget(self.pcvd)
        hboxa1.addWidget(aldp)
        hboxa1.addWidget(self.aldp)
        hboxa1.addWidget(tcvd)
        hboxa1.addWidget(self.tcvd)
        layout.addLayout(hboxa1)
        
        layout2.addWidget(self.button)
        
        self.setLayout(layout)

    def onStateChange(self, state):
        if state == Qt.Unchecked:
            if self.sender() == self.select_all_comp:
                self.amat.setChecked(True)
                self.aixtron.setChecked(True)
                self.asm.setChecked(True)
                self.eugene.setChecked(True)
                self.hitachi.setChecked(True)
                self.lam.setChecked(True)
                self.veeco.setChecked(True)
                self.wonik.setChecked(True)
                self.tes.setChecked(True)
                self.orbotech.setChecked(True)
                self.ultratech.setChecked(True)
                self.jusung.setChecked(True)
                self.tokyo.setChecked(True)
                self.nlpcvd.setChecked(True)
                self.aldp.setChecked(True)
                self.pcvd.setChecked(True)
                self.tcvd.setChecked(True)

        elif state == Qt.Checked:
            if self.sender() == self.select_all_comp:
                self.amat.setChecked(False)
                self.aixtron.setChecked(False)
                self.asm.setChecked(False)
                self.eugene.setChecked(False)
                self.hitachi.setChecked(False)
                self.lam.setChecked(False)
                self.veeco.setChecked(False)
                self.wonik.setChecked(False)
                self.tes.setChecked(False)
                self.ultratech.setChecked(False)
                self.jusung.setChecked(False)
                self.tokyo.setChecked(False)
                self.orbotech.setChecked(False)
                self.nlpcvd.setChecked(False)
                self.aldp.setChecked(False)
                self.pcvd.setChecked(False)
                self.tcvd.setChecked(False)

    def plot(self):
        
        ''' plot some random stuff '''
        # random data

        # create an axis
        ax = self.figure.add_subplot(111)
        
        # discards the old graph
        ax.clear()

        # plot data
        x1 = pd.ExcelFile("Competitor-Flat-File.xlsx")
        x1.sheet_names
        df = x1.parse("Revised")
        
        df['Yearly'].fillna(0, inplace = True)
        
        df = df[['Competitor','Equipment','Yearly', 'CY', 'Level 2']]
        nlpcvd_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
        nlpcvd_13_1, nlpcvd_14_1, nlpcvd_15_1, nlpcvd_16_1, nlpcvd_17_1, nlpcvd_18_1 = [[],[], [], [], [], []]
        
        pcvd_13_1, pcvd_14_1, pcvd_15_1, pcvd_16_1, pcvd_17_1, pcvd_18_1 = [[],[], [], [], [], []]
        tcvd_13_1, tcvd_14_1, tcvd_15_1, tcvd_16_1, tcvd_17_1, tcvd_18_1 = [[],[], [], [], [], []]
        aldp_13_1, aldp_14_1, aldp_15_1, aldp_16_1, aldp_17_1, aldp_18_1 = [[],[], [], [], [], []]
        for year in nlpcvd_year:
            if year == 'CY13':
                nlpcvd = df[(df.Yearly>0) & (df.CY == 'CY13') & (df['Level 2'] ==  'Nontube LPCVD')]
                nlpcvd = sum(nlpcvd['Yearly'])
                nlpcvd_13_1.append(nlpcvd)
                pcvd = df[(df.Yearly>0) & (df.CY == 'CY13') & (df['Level 2'] ==  'Plasma CVD')]
                pcvd = sum(pcvd['Yearly'])
                pcvd_13_1.append(pcvd)
                tcvd = df[(df.Yearly>0) & (df.CY == 'CY13') & (df['Level 2'] ==  'Tube CVD')]
                tcvd = sum(tcvd['Yearly'])
                tcvd_13_1.append(tcvd)
                aldp = df[(df.Yearly>0) & (df.CY == 'CY13') & (df['Level 2'] ==  'Atomic Layer Deposition Platforms')]
                aldp = sum(aldp['Yearly'])
                aldp_13_1.append(aldp)
            if year == 'CY14':
                nlpcvd = df[(df.Yearly>0) & (df.CY == 'CY14') & (df['Level 2'] ==  'Nontube LPCVD')]
                nlpcvd = sum(nlpcvd['Yearly'])
                nlpcvd_14_1.append(nlpcvd)
                pcvd = df[(df.Yearly>0) & (df.CY == 'CY14') & (df['Level 2'] ==  'Plasma CVD')]
                pcvd = sum(pcvd['Yearly'])
                pcvd_14_1.append(pcvd)
                tcvd = df[(df.Yearly>0) & (df.CY == 'CY14') & (df['Level 2'] ==  'Tube CVD')]
                tcvd = sum(tcvd['Yearly'])
                tcvd_14_1.append(tcvd)
                aldp = df[(df.Yearly>0) & (df.CY == 'CY14') & (df['Level 2'] ==  'Atomic Layer Deposition Platforms')]
                aldp = sum(aldp['Yearly'])
                aldp_14_1.append(aldp)
            if year == 'CY15':
                nlpcvd = df[(df.Yearly>0) & (df.CY == 'CY15') & (df['Level 2'] ==  'Nontube LPCVD')]
                nlpcvd = sum(nlpcvd['Yearly'])
                nlpcvd_15_1.append(nlpcvd)
                pcvd = df[(df.Yearly>0) & (df.CY == 'CY15') & (df['Level 2'] ==  'Plasma CVD')]
                pcvd = sum(pcvd['Yearly'])
                pcvd_15_1.append(pcvd)
                tcvd = df[(df.Yearly>0) & (df.CY == 'CY15') & (df['Level 2'] ==  'Tube CVD')]
                tcvd = sum(tcvd['Yearly'])
                tcvd_15_1.append(tcvd)
                aldp = df[(df.Yearly>0) & (df.CY == 'CY15') & (df['Level 2'] ==  'Atomic Layer Deposition Platforms')]
                aldp = sum(aldp['Yearly'])
                aldp_15_1.append(aldp)
                
            if year == 'CY16':
                nlpcvd = df[(df.Yearly>0) & (df.CY == 'CY16') & (df['Level 2'] ==  'Nontube LPCVD')]
                nlpcvd = sum(nlpcvd['Yearly'])
                nlpcvd_16_1.append(nlpcvd)
                pcvd = df[(df.Yearly>0) & (df.CY == 'CY16') & (df['Level 2'] ==  'Plasma CVD')]
                pcvd = sum(pcvd['Yearly'])
                pcvd_16_1.append(pcvd)
                tcvd = df[(df.Yearly>0) & (df.CY == 'CY16') & (df['Level 2'] ==  'Tube CVD')]
                tcvd = sum(tcvd['Yearly'])
                tcvd_16_1.append(tcvd)
                aldp = df[(df.Yearly>0) & (df.CY == 'CY16') & (df['Level 2'] ==  'Atomic Layer Deposition Platforms')]
                aldp = sum(aldp['Yearly'])
                aldp_16_1.append(aldp)
                
            if year == 'CY17':
                nlpcvd = df[(df.Yearly>0) & (df.CY == 'CY17') & (df['Level 2'] ==  'Nontube LPCVD')]
                nlpcvd = sum(nlpcvd['Yearly'])
                nlpcvd_17_1.append(nlpcvd)
                pcvd = df[(df.Yearly>0) & (df.CY == 'CY17') & (df['Level 2'] ==  'Plasma CVD')]
                pcvd = sum(pcvd['Yearly'])
                pcvd_17_1.append(pcvd)
                tcvd = df[(df.Yearly>0) & (df.CY == 'CY17') & (df['Level 2'] ==  'Tube CVD')]
                tcvd = sum(tcvd['Yearly'])
                tcvd_17_1.append(tcvd)
                aldp = df[(df.Yearly>0) & (df.CY == 'CY17') & (df['Level 2'] ==  'Atomic Layer Deposition Platforms')]
                aldp = sum(aldp['Yearly'])
                aldp_17_1.append(aldp)
                
            if year == 'CY18':
                nlpcvd = df[(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  'Nontube LPCVD')]
                nlpcvd = sum(nlpcvd['Yearly'])
                nlpcvd_18_1.append(nlpcvd)
                pcvd = df[(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  'Plasma CVD')]
                pcvd = sum(pcvd['Yearly'])
                pcvd_18_1.append(pcvd)
                tcvd = df[(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  'Tube CVD')]
                tcvd = sum(tcvd['Yearly'])
                tcvd_18_1.append(tcvd)
                aldp = df[(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  'Atomic Layer Deposition Platforms')]
                aldp = sum(aldp['Yearly'])
                aldp_18_1.append(aldp)
        
        nlpcvd_13, pcvd_13,tcvd_13, aldp_13 = [round(sum(nlpcvd_13_1),0),round(sum(pcvd_13_1),0),round(sum(tcvd_13_1),0),round(sum(aldp_13_1),0) ]
        nlpcvd_14,pcvd_14,tcvd_14,aldp_14 = [round(sum(nlpcvd_14_1),0),round(sum(pcvd_14_1),0),round(sum(tcvd_14_1),0),round(sum(aldp_14_1),0)]
        nlpcvd_15, pcvd_15,tcvd_15,aldp_15 = [round(sum(nlpcvd_15_1),0),round(sum(pcvd_15_1),0),round(sum(tcvd_15_1),0),round(sum(aldp_15_1),0)]
        nlpcvd_16,pcvd_16,tcvd_16,aldp_16 = [round(sum(nlpcvd_16_1),0),round(sum(pcvd_16_1),0),round(sum(tcvd_16_1),0),round(sum(aldp_16_1),0)]
        nlpcvd_17,pcvd_17,tcvd_17,aldp_17 = [round(sum(nlpcvd_17_1),0),round(sum(pcvd_17_1),0),round(sum(tcvd_17_1),0),round(sum(aldp_17_1),0)]
        nlpcvd_18,pcvd_18,tcvd_18,aldp_18 = [round(sum(nlpcvd_18_1),0),round(sum(pcvd_18_1),0),round(sum(tcvd_18_1),0),round(sum(aldp_18_1),0)]
        nlpcvd_all = [nlpcvd_13, nlpcvd_14, nlpcvd_15, nlpcvd_16, nlpcvd_17, nlpcvd_18]
        pcvd_all = [pcvd_13, pcvd_14, pcvd_15, pcvd_16, pcvd_17, pcvd_18]
        tcvd_all = [tcvd_13, tcvd_14, tcvd_15, tcvd_16, tcvd_17, tcvd_18]
        aldp_all = [aldp_13, aldp_14, aldp_15, aldp_16, aldp_17, aldp_18]
        nlpcvd_pcvd_aldp_all = [sum([nlpcvd_13,pcvd_13, aldp_13]),sum([nlpcvd_14,pcvd_14, aldp_14]),sum([nlpcvd_15,pcvd_15, aldp_15]),sum([nlpcvd_16,pcvd_16, aldp_16]),sum([nlpcvd_17,pcvd_17, aldp_17]),sum([nlpcvd_18,pcvd_18, aldp_18])]
        nlpcvd_pcvd_tcvd_all = [sum([nlpcvd_13,pcvd_13, tcvd_13]),sum([nlpcvd_14,pcvd_14, tcvd_14]),sum([nlpcvd_15,pcvd_15, tcvd_15]),sum([nlpcvd_16,pcvd_16, tcvd_16]),sum([nlpcvd_17,pcvd_17, tcvd_17]),sum([nlpcvd_18,pcvd_18, tcvd_18])]
        tcvd_aldp_pcvd_all = [sum([tcvd_13,pcvd_13, aldp_13]),sum([tcvd_14,pcvd_14, aldp_14]),sum([tcvd_15,pcvd_15, aldp_15]),sum([tcvd_16,pcvd_16, aldp_16]),sum([tcvd_17,pcvd_17, aldp_17]),sum([tcvd_18,pcvd_18, aldp_18])]
        nlpcvd_pcvd_all = [sum([nlpcvd_13,pcvd_13]),sum([nlpcvd_14,pcvd_14]),sum([nlpcvd_15,pcvd_15]),sum([nlpcvd_16,pcvd_16]),sum([nlpcvd_17,pcvd_17]),sum([nlpcvd_18,pcvd_18])]
        nlpcvd_aldp_all = [sum([nlpcvd_13, aldp_13]),sum([nlpcvd_14, aldp_14]),sum([nlpcvd_15, aldp_15]),sum([nlpcvd_16, aldp_16]),sum([nlpcvd_17, aldp_17]),sum([nlpcvd_18, aldp_18])]
        tcvd_pcvd_all = [sum([tcvd_13,pcvd_13]),sum([tcvd_14,pcvd_14]),sum([tcvd_15,pcvd_15]),sum([tcvd_16,pcvd_16]),sum([tcvd_17,pcvd_17]),sum([tcvd_18,pcvd_18])]
        tcvd_aldp_all = [sum([tcvd_13,aldp_13]),sum([tcvd_14,aldp_14]),sum([tcvd_15,aldp_15]),sum([tcvd_16,aldp_16]),sum([tcvd_17,aldp_17]),sum([tcvd_18,aldp_18])]
        nlpcvd_tcvd_all  = [sum([nlpcvd_13, tcvd_13]),sum([nlpcvd_14, tcvd_14]),sum([nlpcvd_15, tcvd_15]),sum([nlpcvd_16, tcvd_16]),sum([nlpcvd_17, tcvd_17]),sum([nlpcvd_18, tcvd_18])]
        aldp_pcvd_all = [sum([aldp_13,pcvd_13]),sum([aldp_14,pcvd_14]),sum([aldp_15,pcvd_15]),sum([aldp_16,pcvd_16]),sum([aldp_17,pcvd_17]),sum([aldp_18,pcvd_18])]
        
        if self.nlpcvd.isChecked() and self.pcvd.isChecked() and self.aldp.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_cal = []
                amat_total_13 =691.3738
                amat_total_14 = 1057.973
                amat_total_15 = 1182.857
                amat_total_16 = 1782.115
                amat_total_17 = 2426.448
                
                amat_eq = ['Atomic Layer Deposition Platforms', 'Nontube LPCVD', 'Plasma CVD']
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 0)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/nlpcvd_pcvd_aldp_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
            
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[blue_patch])
                
            if self.lam.isChecked():
                lam = []
                lam_cal = []
                lam_total_13 = 580.4183
                lam_total_14 = 805.3354
                lam_total_15 = 901.35
                lam_total_16 = 1151.953
                lam_total_17 = 2184.25
                
                lam_eq = ['Atomic Layer Deposition Platforms', 'Nontube LPCVD', 'Plasma CVD']
                            
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                    
                lam_total_18 = round(sum(lam), 0)
                
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/nlpcvd_pcvd_aldp_all[num])*100, 2))
            
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                if self.amat.isChecked():
                    red_patch = mpatches.Patch(color = 'red', label = 'LAM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[red_patch,blue_patch])
                else:
                    red_patch = mpatches.Patch(color = 'red', label = 'LAM')
                    ax.legend(handles=[red_patch])
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 107.9
                wonik_total_14 = 124.0
                wonik_total_15 = 128.4769
                wonik_total_16 = 137.2021
                wonik_total_17 = 372.1
                wonik_cal = []
                            
                wonik_eq = ['Atomic Layer Deposition Platforms', 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                    
                wonik_total_18 = round(sum(wonik), 0)
                
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]
                
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/nlpcvd_pcvd_aldp_all[num])*100, 2))
            
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                
                if self.amat.isChecked() and self.lam.isChecked():
                    golden_patch = mpatches.Patch(color = 'goldenrod', label = 'Wonik')
                    red_patch = mpatches.Patch(color = 'red', label = 'LAM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[golden_patch,red_patch,blue_patch])
                elif self.amat.isChecked():
                    golden_patch = mpatches.Patch(color = 'goldenrod', label = 'Wonik')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[golden_patch,blue_patch])
                elif self.lam.isChecked():
                    golden_patch = mpatches.Patch(color = 'goldenrod', label = 'Wonik')
                    red_patch = mpatches.Patch(color = 'red', label = 'LAM')
                    ax.legend(handles=[golden_patch,red_patch])
                else:
                    golden_patch = mpatches.Patch(color = 'goldenrod', label = 'Wonik')
                    ax.legend(handles=[golden_patch])
        
        elif self.nlpcvd.isChecked() and self.pcvd.isChecked() and self.tcvd.isChecked():
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 406.9015
                tokyo_total_14 = 643.9228
                tokyo_total_15 = 663.3027
                tokyo_total_16 = 583.8111
                tokyo_total_17 = 1014.6
                tokyo_cal = []
                            
                tokyo_eq = ['Nontube LPCVD', 'Plasma CVD','Tube CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/nlpcvd_pcvd_tcvd_all[num])*100, 2))
            
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data', fontsize = 8)
                brown_patch = mpatches.Patch(color='brown',label='Tokyo Electron')
                ax.legend(handles=[brown_patch])
        
        elif self.tcvd.isChecked() and self.aldp.isChecked() and self.pcvd.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 401.5537
                asm_total_14 = 500.5012
                asm_total_15 = 520.3123
                asm_total_16 = 429.0665
                asm_total_17 = 526.7842
                asm_cal = []
                
                asm_eq = ['Atomic Layer Deposition Platforms','Plasma CVD','Tube CVD' ]
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/tcvd_aldp_pcvd_all[num])*100, 2))
    
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)
                    
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    ax.legend(handles=[orange_patch])
            
        elif self.nlpcvd.isChecked() and self.pcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 =691.3738
                amat_total_14 = 1057.973
                amat_total_15 = 1156.762
                amat_total_16 = 1686.915
                amat_total_17 = 2365.264
                amat_all = []
                amat_cal = []
                
                amat_eq = ['Nontube LPCVD', 'Plasma CVD']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 1)
                
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
                
                    
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/nlpcvd_pcvd_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
            
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[blue_patch])
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 58.0
                eugene_total_14 = 35.9
                eugene_total_15 = 42.1
                eugene_total_16 = 38.779
                eugene_total_17 = 48.200
                eugene_all = []
                eugene_cal = []
                
                eugene_eq = ['Nontube LPCVD', 'Plasma CVD']
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = (i['Yearly'])
                    eugene.append(ii)
                    
                eugene_total_18 = round(sum(eugene), 1)
                
                eugene_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
                
                    
                for num in [0,1,2,3,4,5]:
                    eugene_cal.append(round((eugene_total[num]/nlpcvd_pcvd_all[num])*100, 2))
            
                ax.scatter(eugene_year, eugene_cal, color = 'green')
                ax.plot(eugene_year, eugene_cal, color ='green')
                
                ax.set_xticks(eugene_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % eugene_cal[num], xy = (eugene_year[num],eugene_cal[num]), textcoords='data', fontsize = 8)
            
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color = 'green', label = 'Eugene')
                    
                    ax.legend(handles=[green_patch,blue_patch])
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 580.4183
                lam_total_14 = 759.7654
                lam_total_15 = 805.95
                lam_total_16 = 1088.365
                lam_total_17 = 1969.35
                lam_cal = []
                
                lam_eq = [ 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/nlpcvd_aldp_all[num])*100, 2))
            
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                    
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color = 'green', label = 'Eugene')
                    
                    ax.legend(handles=[red_patch,green_patch,blue_patch])
            if self.tes.isChecked():
                tes = []
                tes_total_13 = 28.0
                tes_total_14 = 72.6
                tes_total_15 = 68.4
                tes_total_16 = 115.301
                tes_total_17 = 180.6
                tes_cal = []
                
                tes_eq = [ 'Nontube LPCVD', 'Plasma CVD']
                tes_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in tes_eq:
                    i = df[(df.Competitor == 'Tes' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tes.append(ii)
                
                tes_total_18 = round(sum(tes),1)
                
                tes_total = [tes_total_13,tes_total_14, tes_total_15, tes_total_16, tes_total_17, tes_total_18]
                
                for num in [0,1,2,3,4,5]:
                    tes_cal.append(round((tes_total[num]/nlpcvd_pcvd_all[num])*100, 2))
                
                ax.scatter(tes_year, tes_cal, color = 'magenta')
                ax.plot(tes_year, tes_cal, color ='magenta')
                
                ax.set_xticks(tes_year)
                
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tes_cal[num], xy = (tes_year[num],tes_cal[num]), textcoords='data', fontsize = 8)
                    
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color = 'green', label = 'Eugene')
                    
                    ax.legend(handles=[magenta_patch,red_patch,green_patch,blue_patch])
                    
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 104.4687
                tokyo_total_14 = 204.9875
                tokyo_total_15 = 246.966
                tokyo_total_16 = 237.7651
                tokyo_total_17 = 439.3
                tokyo_cal = []
                
                tokyo_eq = ['Nontube LPCVD', 'Plasma CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),1)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
                
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                                    
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/nlpcvd_pcvd_all[num])*100, 2))
                    
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data' ,fontsize = 8)

                    brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color = 'green', label = 'Eugene')
                    
                    ax.legend(handles=[brown_patch,magenta_patch,red_patch,green_patch,blue_patch])
                    
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 68.5
                wonik_total_14 = 78.7
                wonik_total_15 = 80.7259
                wonik_total_16 = 91.5999
                wonik_total_17 = 248.4
                wonik_cal = []
                
                wonik_eq = [ 'Nontube LPCVD', 'Plasma CVD']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),1)
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
                
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
                
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/nlpcvd_pcvd_all[num])*100, 2))
                    
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                    
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik')
                    brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color = 'green', label = 'Eugene')
                    
                    ax.legend(handles=[golden_patch,brown_patch,magenta_patch,red_patch,green_patch,blue_patch])
                
        elif self.nlpcvd.isChecked() and self.aldp.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 =72.141341
                amat_total_14 = 133.818248
                amat_total_15 = 114.9833
                amat_total_16 = 343.8527
                amat_total_17 = 209.4539
                amat_cal = []
                amat_all = []
                
                amat_eq = ['Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 0)
                
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
                
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/nlpcvd_aldp_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
            
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[blue_patch])
                    
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 125.6684703
                lam_total_14 = 224.2045
                lam_total_15 = 303.75
                lam_total_16 = 283.3968
                lam_total_17 = 735.1
                lam_cal = []
                lam_all = []
                
                lam_eq = [ 'Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/nlpcvd_aldp_all[num])*100, 2))
            
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                    
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[red_patch,blue_patch])
                    
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 39.4
                wonik_total_14 = 45.3
                wonik_total_15 = 47.75091049
                wonik_total_16 = 75.60211945
                wonik_total_17 = 174.9
                wonik_all = []
                wonik_cal = []
                
                wonik_eq = [ 'Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]

                
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/nlpcvd_aldp_all[num])*100, 2))
                    
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                    
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[golden_patch,red_patch,blue_patch])
                    
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 44.1
                eugene_total_14 = 27.3
                eugene_total_15 = 29.9
                eugene_total_16 = 38.77942
                eugene_total_17 = 60.5
                eugene_cal = []
                
                eugene_eq = [ 'Nontube LPCVD', 'Atomic Layer Deposition Platforms']
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                
                eugene_total_18 = round(sum(eugene),0)
                                
                eugene_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
                
                for num in [0,1,2,3,4,5]:
                    eugene_cal.append(round((eugene_total[num]/nlpcvd_aldp_all[num])*100, 2))
            
                ax.scatter(eugene_year, eugene_cal, color = 'green')
                ax.plot(eugene_year, eugene_cal, color ='green')
                
                ax.set_xticks(eugene_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % eugene_cal[num], xy = (eugene_year[num],eugene_cal[num]), textcoords='data', fontsize = 8)

                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[green_patch,golden_patch,red_patch,blue_patch])
        
        elif self.tcvd.isChecked() and self.pcvd.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 53.5832
                asm_total_14 = 54.3012
                asm_total_15 = 51.8123
                asm_total_16 = 39.386
                asm_total_17 = 80.6
                asm_cal = []
                
                asm_eq = ['Tube CVD', 'Plasma CVD']
                
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16,asm_total_17, asm_total_18]
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/tcvd_pcvd_all[num])*100, 2))
    
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)
                    if self.tokyo.isChecked():
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        orange_patch = mpatches.Patch(color='orange',label='ASM')
                        ax.legend(handles=[orange_patch, brown_patch])
                    else:
                        orange_patch = mpatches.Patch(color='orange',label='ASM')
                        ax.legend(handles=[orange_patch])
                    
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 302.432679
                tokyo_total_14 = 438.935
                tokyo_total_15 = 416.3365
                tokyo_total_16 = 350.2369
                tokyo_total_17 = 583.5
                tokyo_cal = []
                
                tokyo_eq = ['Tube CVD', 'Plasma CVD' ]
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
                
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/tcvd_pcvd_all[num])*100, 2))
                    
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data' ,fontsize = 8)
                    
                    if self.asm.isChecked():
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        orange_patch = mpatches.Patch(color='orange',label='ASM')
                        ax.legend(handles=[orange_patch, brown_patch])
                    else:
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        ax.legend(handles=[ brown_patch]) 
                        
        elif self.tcvd.isChecked() and self.aldp.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 371.2758
                asm_total_14 = 469.79
                asm_total_15 = 495.8
                asm_total_16 = 411.959
                asm_total_17 = 473.5842
                asm_cal = []
                
                asm_eq = ['Tube CVD', 'Atomic Layer Deposition Platforms']
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                    
                asm_total_18 = round(sum(asm), 1)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16,asm_total_17, asm_total_18]
                
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/tcvd_aldp_all[num])*100, 2))
    
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)

                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    ax.legend(handles=[orange_patch])
                    
        elif self.nlpcvd.isChecked() and self.tcvd.isChecked():
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 406.9015
                tokyo_total_14 = 643.9228
                tokyo_total_15 = 663.3027
                tokyo_total_16 = 579.6202
                tokyo_total_17 = 1006.4
                tokyo_cal = []
                
                tokyo_eq = ['Tube CVD', 'Nontube LPCVD' ]
                
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
        
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/nlpcvd_tcvd_all[num])*100, 2))
                    
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data' ,fontsize = 8)

                    brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                    ax.legend(handles=[ brown_patch])
                    
        elif self.aldp.isChecked() and self.pcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 619.232479
                amat_total_14 = 924.1551212
                amat_total_15 = 1093.968
                amat_total_16 = 1533.462
                amat_total_17 = 2278.178
                amat_cal = []
                
                amat_eq = ['Atomic Layer Deposition Platforms', 'Plasma CVD']
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                    
                amat_total_18 = round(sum(amat), 1)
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
                
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/aldp_pcvd_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
            
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[blue_patch])
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 378.2477
                asm_total_14 = 476.9612
                asm_total_15 = 493.0123
                asm_total_16 = 406.788
                asm_total_17 = 499.3842
                asm_cal = []
                
                asm_eq = ['Atomic Layer Deposition Platforms', 'Plasma CVD']
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                    
                asm_total_18 = round(sum(asm), 1)
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16,asm_total_17, asm_total_18]
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/aldp_pcvd_all[num])*100, 2))
            
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)
                    
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[orange_patch,blue_patch])
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 580.4183
                lam_total_14 = 759.7654
                lam_total_15 = 805.95
                lam_total_16 = 1088.365
                lam_total_17 = 1969.35
                lam_cal= []
                
                lam_eq = [ 'Atomic Layer Deposition Platforms', 'Plasma CVD']
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),1)
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/aldp_pcvd_all[num])*100, 2))
            
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                    
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[red_patch,orange_patch,blue_patch])
            
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 107.9
                wonik_total_14 = 124.0
                wonik_total_15 = 128.4769
                wonik_total_16 = 107.2021
                wonik_total_17 = 320.7
                wonik_cal = []
                
                wonik_eq = [ 'Atomic Layer Deposition Platforms', 'Plasma CVD']
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),1)
                
                wonik_total = [wonik_total_13,wonik_total_14, wonik_total_15,wonik_total_16, wonik_total_17, wonik_total_18]
            
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/aldp_pcvd_all[num])*100, 2))
                    
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                    
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[golden_patch,red_patch,orange_patch,blue_patch])
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 13.9
                eugene_total_14 = 8.6
                eugene_total_15 = 12.2
                eugene_total_16 = 0
                eugene_total_17 = 12.300
                eugene_cal = []
                
                eugene_eq = [ 'Plasma CVD', 'Atomic Layer Deposition Platforms']
                eugene_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii =(i['Yearly'])
                    eugene.append(ii)
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_total = [eugene_total_13,eugene_total_14, eugene_total_15,eugene_total_16, eugene_total_17, eugene_total_18]
                
                for num in [0,1,2,3,4,5]:
                    eugene_cal.append(round((eugene_total[num]/aldp_pcvd_all[num])*100, 2))
            
                ax.scatter(eugene_year, eugene_cal, color = 'green')
                ax.plot(eugene_year, eugene_cal, color ='green')
                
                ax.set_xticks(eugene_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % eugene_cal[num], xy = (eugene_year[num],eugene_cal[num]), textcoords='data', fontsize = 8)

                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    
                    ax.legend(handles=[green_patch,golden_patch,red_patch,orange_patch,blue_patch])
        
        elif self.tcvd.isChecked():
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 23.30597
                asm_total_14 = 23.54
                asm_total_15 = 27.3
                asm_total_16 = 22.278499
                asm_total_17 = 27.399999
                
                asm_eq = ['Tube CVD' ]
                asm_cal = []
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                    
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/tcvd_all[num])*100, 2))
            
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)
                    
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    ax.legend(handles=[orange_patch])
                
            if self.hitachi.isChecked():
                hitachi = []
                hitachi_total_13 = 318.3580273
                hitachi_total_14 = 479.1185744
                hitachi_total_15 = 504.640823
                hitachi_total_16 = 412.2118835
                hitachi_total_17 = 776.1200562
                
                hitachi_eq = ['Tube CVD' ]
                hitachi_all = []
                hitachi_cal = []
                
                for eq in hitachi_eq:
                    i = df[(df.Competitor == 'Hitachi Kokusai Electric' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    hitachi.append(ii)
                
                hitachi_total_18 = round(sum(hitachi),0)
                
                hitachi_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                hitachi_total = [hitachi_total_13, hitachi_total_14, hitachi_total_15, hitachi_total_16, hitachi_total_17, hitachi_total_18]
                
                    
                for num in [0,1,2,3,4,5]:
                    hitachi_cal.append(round((hitachi_total[num]/tcvd_all[num])*100, 2))
                    
                ax.scatter(hitachi_year, hitachi_cal, color = 'yellow')
                ax.plot(hitachi_year, hitachi_cal, color ='yellow')
                
                ax.set_xticks(hitachi_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % hitachi_cal[num], xy = (hitachi_year[num],hitachi_cal[num]), textcoords='data', fontsize = 8)
                    
                    if self.asm.isChecked():
                        orange_patch = mpatches.Patch(color='orange',label='ASM')
                        yellow_patch = mpatches.Patch(color='yellow',label='Hitachi')
                        ax.legend(handles=[orange_patch,yellow_patch])
                    else:
                        yellow_patch = mpatches.Patch(color='yellow',label='Hitachi')
                        ax.legend(handles=[yellow_patch])
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 302.4326797
                tokyo_total_14 = 438.9352201
                tokyo_total_15 = 416.3365286
                tokyo_total_16 = 346.0460205
                tokyo_total_17 = 575.2999878
                
                tokyo_eq = ['Tube CVD' ]
                tokyo_all = []
                tokyo_cal = []
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
                    
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/tcvd_all[num])*100, 2))
                    
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data' ,fontsize = 8)
                    
                    if self.asm.isChecked() and self.hitachi.isChecked():
                        orange_patch = mpatches.Patch(color='orange',label='ASM')
                        yellow_patch = mpatches.Patch(color='yellow',label='Hitachi')
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        ax.legend(handles=[brown_patch,orange_patch,yellow_patch])
                    elif self.asm.isChecked():
                        orange_patch = mpatches.Patch(color='orange',label='ASM')
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        ax.legend(handles=[brown_patch,orange_patch])
                    elif self.hitachi.isChecked():
                        yellow_patch = mpatches.Patch(color='yellow',label='Hitachi')
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        ax.legend(handles=[brown_patch,yellow_patch])
                    else:
                        brown_patch = mpatches.Patch(color='brown',label='Tokyo')
                        ax.legend(handles=[brown_patch])
        elif self.nlpcvd.isChecked():
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 72.1413
                amat_total_14 = 133.818
                amat_total_15 = 88.889
                amat_total_16 = 248.652
                amat_total_17 = 148.270
                amat_cal = []
                amat_all = []
                amat_eq = ['Nontube LPCVD' ]
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                
                amat_total_18 = round(sum(amat),0)
                
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
            
                
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/nlpcvd_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())

            
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
                
                blue_patch = mpatches.Patch(color='blue',label='AMAT')
                ax.legend(handles=[blue_patch])
            if self.eugene.isChecked():
                
                eugene = []
                eugene_total_13 = 44.1
                eugene_total_14 = 27.3
                eugene_total_15 = 29.9
                eugene_total_16 = 38.7794
                eugene_total_17 = 48.200
                eugene_all = []
                eugene_cal = []
                
                eugene_eq = ['Nontube LPCVD' ]
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                    
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
        
                
                for num in [0,1,2,3,4,5]:
                    eugene_cal.append(round((eugene_total[num]/nlpcvd_all[num])*100, 2))
            
                ax.scatter(eugene_year, eugene_cal, color = 'green')
                ax.plot(eugene_year, eugene_cal, color ='green')
                
                ax.set_xticks(eugene_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % eugene_cal[num], xy = (eugene_year[num],eugene_cal[num]), textcoords='data', fontsize = 8)
                    if self.amat.isChecked():
                        blue_patch = mpatches.Patch(color='blue',label='AMAT')
                        green_patch = mpatches.Patch(color='green',label='Eugene')
                        ax.legend(handles=[blue_patch,green_patch])
                    else:
                        green_patch = mpatches.Patch(color='green',label='Eugene')
                        ax.legend(handles=[green_patch])
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 125.6684
                lam_total_14 = 178.6345
                lam_total_15 = 208.35
                lam_total_16 = 219.8092499
                lam_total_17 = 520.0999
                lam_cal = []
                
                lam_eq = ['Nontube LPCVD' ]
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                lam_total = [lam_total_13, lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/nlpcvd_all[num])*100, 2))
                
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
                
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                    
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    red_patch = mpatches.Patch(color='red',label='Lam Research')
                    ax.legend(handles=[blue_patch, green_patch, red_patch])
            if self.tes.isChecked():
                tes = []
                tes_total_13 = 0.0
                tes_total_14 = 0.0
                tes_total_15 = 0.0
                tes_total_16 = 40.001
                tes_total_17 = 76.400
                tes_all = []
                tes_cal = []
                
                tes_eq = ['Nontube LPCVD' ]
                
                for eq in tes_eq:
                    i = df[(df.Competitor == 'Tes' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tes.append(ii)
                
                tes_total_18 = round(sum(tes),0)
                
                tes_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                tes_total = [tes_total_13, tes_total_14, tes_total_15, tes_total_16, tes_total_17, tes_total_18]
                
                for num in [0,1,2,3,4,5]:
                    tes_cal.append(round((tes_total[num]/nlpcvd_all[num])*100, 2))
                
                ax.scatter(tes_year, tes_cal, color = 'magenta')
                ax.plot(tes_year, tes_cal, color ='magenta')
                
                ax.set_xticks(tes_year)
                
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tes_cal[num], xy = (tes_year[num],tes_cal[num]), textcoords='data', fontsize = 8)
                    
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    red_patch = mpatches.Patch(color='red',label='Lam Research')
                    ax.legend(handles=[blue_patch, green_patch, red_patch, magenta_patch])
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 104.4687
                tokyo_total_14 = 204.98759
                tokyo_total_15 = 246.966132
                tokyo_total_16 = 233.5742035
                tokyo_total_17 = 431.1000
                tokyo_all = []
                tokyo_cal = []
                
                tokyo_eq = ['Nontube LPCVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/nlpcvd_all[num])*100, 2))
            
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data', fontsize = 8)
                brown_patch = mpatches.Patch(color='brown',label='Tokyo Electron')
                magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                blue_patch = mpatches.Patch(color='blue',label='AMAT')
                green_patch = mpatches.Patch(color='green',label='Eugene')
                red_patch = mpatches.Patch(color='red',label='Lam Research')
                ax.legend(handles=[blue_patch, green_patch, red_patch, magenta_patch, brown_patch])
            
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 0.0
                wonik_total_14 = 0.0
                wonik_total_15 = 0.0
                wonik_total_16 = 30.0
                wonik_total_17 = 51.2999
                wonik_all= []
                wonik_cal = []
                
                wonik_eq = ['Nontube LPCVD' ]
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                wonik_total = [wonik_total_13, wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]
                

                
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/nlpcvd_all[num])*100, 2))
            
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                
                golden_patch = mpatches.Patch(color='goldenrod',label='Wonik IPS')
                brown_patch = mpatches.Patch(color='brown',label='Tokyo Electron')
                magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                blue_patch = mpatches.Patch(color='blue',label='AMAT')
                green_patch = mpatches.Patch(color='green',label='Eugene')
                red_patch = mpatches.Patch(color='red',label='Lam Research')
                ax.legend(handles=[blue_patch, green_patch, red_patch, magenta_patch, brown_patch, golden_patch])
        
            if self.aixtron.isChecked():
                aixtron = []
                aixtron_total_13 = 0.0
                aixtron_total_14 = 0.0
                aixtron_total_15 = 0.0
                aixtron_total_16 = 0.0
                aixtron_total_17 = 0.8999
                aixtron_all = []
                aixtron_cal = []
                
                aixtron_eq = ['Nontube LPCVD' ]
                
                for eq in aixtron_eq:
                    i = df[(df.Competitor == 'Aixtron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    aixtron.append(ii)
                
                aixtron_total_18 = round(sum(aixtron),1)
                
                aixtron_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                aixtron_total = [aixtron_total_13, aixtron_total_14, aixtron_total_15, aixtron_total_16, aixtron_total_17, aixtron_total_18]
            
                
                for num in [0,1,2,3,4,5]:
                    aixtron_cal.append(round((aixtron_total[num]/nlpcvd_all[num])*100, 2))
            
                ax.scatter(aixtron_year, aixtron_cal, color = 'lightgrey')
                ax.plot(aixtron_year, aixtron_cal, color ='lightgrey')
                
                ax.set_xticks(aixtron_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % aixtron_cal[num], xy = (aixtron_year[num],aixtron_cal[num]), textcoords='data', fontsize = 8)
            
                grey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                golden_patch = mpatches.Patch(color='goldenrod',label='Wonik IPS')
                brown_patch = mpatches.Patch(color='brown',label='Tokyo Electron')
                magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                blue_patch = mpatches.Patch(color='blue',label='AMAT')
                green_patch = mpatches.Patch(color='green',label='Eugene')
                red_patch = mpatches.Patch(color='red',label='Lam Research')
                ax.legend(handles=[grey_patch, blue_patch, green_patch, red_patch, magenta_patch, brown_patch, golden_patch])
            if self.hitachi.isChecked():
                
                hitachi_total_13 = 0.0
                hitachi_total_14 = 0.0
                hitachi_total_15 = 2.925022
                hitachi_total_16 = 0.0
                hitachi_total_17 = 0.0
                hitachi_total_18 = 0.0
                hitachi_all = []
                hitachi_cal = []
                
                hitachi_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                hitachi_total = [hitachi_total_13, hitachi_total_14, hitachi_total_15, hitachi_total_16, hitachi_total_17, hitachi_total_18]
            
                for num in [0,1,2,3,4,5]:
                    hitachi_cal.append(round((hitachi_total[num]/nlpcvd_all[num])*100, 2))
            
                ax.scatter(hitachi_year, hitachi_cal, color = 'yellow')
                ax.plot(hitachi_year, hitachi_cal, color ='yellow')
                
                ax.set_xticks(hitachi_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % hitachi_cal[num], xy = (hitachi_year[num],hitachi_cal[num]), textcoords='data', fontsize = 8)
            
                aqua_patch = mpatches.Patch(color='yellow',label='Hitachi')
                grey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                golden_patch = mpatches.Patch(color='goldenrod',label='Wonik IPS')
                brown_patch = mpatches.Patch(color='brown',label='Tokyo Electron')
                magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                blue_patch = mpatches.Patch(color='blue',label='AMAT')
                green_patch = mpatches.Patch(color='green',label='Eugene')
                red_patch = mpatches.Patch(color='red',label='Lam Research')
                ax.legend(handles=[aqua_patch,grey_patch, blue_patch, green_patch, red_patch, magenta_patch, brown_patch, golden_patch]) 
        
        elif self.pcvd.isChecked(): 
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 619.232479
                amat_total_14 = 924.155121
                amat_total_15 = 1067.973379
                amat_total_16 = 1438.2618
                amat_total_17 = 2216.994
                amat_cal = []
                amat_all = []
                
                amat_eq = ['Plasma CVD' ]
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                
                amat_total_18 = round(sum(amat),0)
                
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]
                
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
                    
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[blue_patch])
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 30.2778
                asm_total_14 = 30.7612
                asm_total_15 = 24.5123
                asm_total_16 = 17.1075
                asm_total_17 = 53.2000
                asm_all = []
                asm_cal = []
                
                asm_eq = ['Plasma CVD' ]
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
            
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)
                    
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[orange_patch,blue_patch])
            if self.eugene.isChecked():
                eugene_total_13 = 13.9
                eugene_total_14 = 8.6
                eugene_total_15 = 12.2
                eugene_total_16 = 0.0
                eugene_total_17 = 0.0
                eugene_total_18 = 0.0
                eugene_cal = []
                eugene_all = []
                
                
                eugene_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]

                
                for num in [0,1,2,3,4,5]:
                    eugene_cal.append(round((eugene_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(eugene_year, eugene_cal, color = 'green')
                ax.plot(eugene_year, eugene_cal, color ='green')
                
                ax.set_xticks(eugene_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % eugene_cal[num], xy = (eugene_year[num],eugene_cal[num]), textcoords='data', fontsize = 8)
                    
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[green_patch,orange_patch,blue_patch])
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 454.749785
                lam_total_14 = 581.1309032
                lam_total_15 = 597.6
                lam_total_16 = 868.5560303
                lam_total_17 = 1449.25
                lam_all = []
                lam_cal = []
                
                lam_eq = ['Plasma CVD' ]
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                lam_total = [lam_total_13, lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                    
                    red_patch = mpatches.Patch(color='red',label='Lam Research')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[red_patch,green_patch,orange_patch,blue_patch])
        
            if self.orbotech.isChecked():
                orbotech = []
                orbotech_total_13 = 0.0
                orbotech_total_14 = 0.0
                orbotech_total_15 = 15.5496
                orbotech_total_16 = 17.0734
                orbotech_total_17 = 23.7000
                orbotech_cal = []
                orbotech_all = []
                
                orbotech_eq = ['Plasma CVD' ]
                
                for eq in orbotech_eq:
                    i = df[(df.Competitor == 'Orbotech' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    orbotech.append(ii)
                
                orbotech_total_18 = round(sum(orbotech),0)
                
                orbotech_year = ['CY15','CY16','CY17', 'CY18']
                
                orbotech_total = [orbotech_total_15, orbotech_total_16, orbotech_total_17, orbotech_total_18]

                
                for num in [0,1,2,3,4,5]:
                    orbotech_cal.append(round((orbotech_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(orbotech_year, orbotech_cal, color = 'lightblue')
                ax.plot(orbotech_year, orbotech_cal, color ='lightblue')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3]:                               # <--
                    ax.annotate('(%s)' % orbotech_cal[num], xy = (orbotech_year[num],orbotech_cal[num]), textcoords='data', fontsize = 8)
                    
                    lightblue_patch = mpatches.Patch(color = 'lightblue', label='Orbotech')
                    red_patch = mpatches.Patch(color='red',label='Lam Research')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[lightblue_patch,red_patch,green_patch,orange_patch,blue_patch])
        
            if self.tes.isChecked():
                tes = []
                tes_total_13 = 28.0
                tes_total_14 = 72.6
                tes_total_15 = 68.4
                tes_total_16 = 75.3000
                tes_total_17 = 104.19999
                tes_cal = []
                tes_all = []
                
                tes_eq = ['Plasma CVD' ]
                
                for eq in tes_eq:
                    i = df[(df.Competitor == 'Tes' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tes.append(ii)
                
                tes_total_18 = round(sum(tes),0)
                
                tes_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                tes_total = [tes_total_13, tes_total_14, tes_total_15, tes_total_16, tes_total_17, tes_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    tes_cal.append(round((tes_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(tes_year, tes_cal, color = 'magenta')
                ax.plot(tes_year, tes_cal, color ='magenta')
                
                ax.set_xticks(tes_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tes_cal[num], xy = (tes_year[num],tes_cal[num]), textcoords='data', fontsize = 8)
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    lightblue_patch = mpatches.Patch(color = 'lightblue', label='Orbotech')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[magenta_patch,lightblue_patch,red_patch,green_patch,blue_patch,orange_patch])
                    
            if self.tokyo.isChecked():
                tokyo = []
                tokyo_total_13 = 0.0
                tokyo_total_14 = 0.0
                tokyo_total_15 = 0.0
                tokyo_total_16 = 4.0
                tokyo_total_17 = 8.0
                tokyo_cal = []
                tokyo_all = []
                
                tokyo_eq = ['Plasma CVD' ]
                
                for eq in tokyo_eq:
                    i = df[(df.Competitor == 'Tokyo Electron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    tokyo.append(ii)
                
                tokyo_total_18 = round(sum(tokyo),0)
                
                tokyo_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                tokyo_total = [tokyo_total_13, tokyo_total_14, tokyo_total_15, tokyo_total_16, tokyo_total_17, tokyo_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    tokyo_cal.append(round((tokyo_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(tokyo_year, tokyo_cal, color = 'brown')
                ax.plot(tokyo_year, tokyo_cal, color ='brown')
                
                ax.set_xticks(tokyo_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % tokyo_cal[num], xy = (tokyo_year[num],tokyo_cal[num]), textcoords='data', fontsize = 8)    
                    
                    brown_patch = mpatches.Patch(color = 'brown', label='Tokyo')
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    lightblue_patch = mpatches.Patch(color = 'lightblue', label='Orbotech')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[brown_patch, magenta_patch,lightblue_patch,red_patch,green_patch,blue_patch,orange_patch])
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 68.5
                wonik_total_14 = 78.7
                wonik_total_15 = 80.7259
                wonik_total_16 = 61.5999
                wonik_total_17 = 197.100
                wonik_cal = []
                wonik_all = []
                
                wonik_eq = ['Plasma CVD' ]
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik .append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                wonik_total = [wonik_total_13, wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]

                
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/pcvd_all[num])*100, 2))
            
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                    
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik IPS')
                    brown_patch = mpatches.Patch(color = 'brown', label='Tokyo')
                    magenta_patch = mpatches.Patch(color='magenta',label='Tes')
                    lightblue_patch = mpatches.Patch(color = 'lightblue', label='Orbotech')
                    green_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[golden_patch, brown_patch, magenta_patch,lightblue_patch,red_patch,green_patch,blue_patch,orange_patch])
        
        elif self.aldp.isChecked():
            if self.aixtron.isChecked():
                aixtron = []
                aixtron_total_13 = 41.2336
                aixtron_total_14 = 20.856
                aixtron_total_15 = 33.0
                aixtron_total_16 = 27.059999
                aixtron_total_17 = 39.7999
                aixtron_all = []
                aixtron_cal = []
                
                aixtron_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in aixtron_eq:
                    i = df[(df.Competitor == 'Aixtron' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    aixtron.append(ii)
                
                aixtron_total_18 = round(sum(aixtron),0)
                
                aixtron_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                aixtron_total = [aixtron_total_13, aixtron_total_14, aixtron_total_15, aixtron_total_16, aixtron_total_17, aixtron_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    aixtron_cal.append(round((aixtron_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(aixtron_year, aixtron_cal, color = 'lightgrey')
                ax.plot(aixtron_year, aixtron_cal, color ='lightgrey')
                
                ax.set_xticks(aixtron_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % aixtron_cal[num], xy = (aixtron_year[num],aixtron_cal[num]), textcoords='data', fontsize = 8)
                    
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    ax.legend(handles=[lightgrey_patch])
                    
            if self.amat.isChecked():
                amat = []
                amat_total_13 = 0.0
                amat_total_14 = 0.0
                amat_total_15 = 26.09425
                amat_total_16 = 95.19999
                amat_total_17 = 61.183925
                amat_cal = []
                amat_all = []
                
                amat_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in amat_eq:
                    i = df[(df.Competitor == 'Applied Materials' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    amat.append(ii)
                
                amat_total_18 = round(sum(amat),0)
                
                amat_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                amat_total = [amat_total_13, amat_total_14, amat_total_15, amat_total_16, amat_total_17, amat_total_18]

                
                for num in [0,1,2,3,4,5]:
                    amat_cal.append(round((amat_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(amat_year, amat_cal, color = 'blue')
                ax.plot(amat_year, amat_cal, color ='blue')
                
                ax.set_xticks(amat_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % amat_cal[num], xy = (amat_year[num],amat_cal[num]), textcoords='data', fontsize = 8)
            
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                            
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    ax.legend(handles=[blue_patch,lightgrey_patch])
        
            if self.asm.isChecked():
                asm = []
                asm_total_13 = 347.9698445
                asm_total_14 = 446.2
                asm_total_15 = 468.5
                asm_total_16 = 389.6805115
                asm_total_17 = 446.1841736
                asm_cal = []
                asm_all = []
                
                asm_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in asm_eq:
                    i = df[(df.Competitor == 'ASM International' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    asm.append(ii)
                
                asm_total_18 = round(sum(asm),0)
                
                asm_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
                
                asm_total = [asm_total_13, asm_total_14, asm_total_15, asm_total_16, asm_total_17, asm_total_18]
                
                for num in [0,1,2,3,4,5]:
                    asm_cal.append(round((asm_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(asm_year, asm_cal, color = 'orange')
                ax.plot(asm_year, asm_cal, color ='orange')
                
                ax.set_xticks(asm_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % asm_cal[num], xy = (asm_year[num],asm_cal[num]), textcoords='data', fontsize = 8)
                    
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[orange_patch,lightgrey_patch,blue_patch])
        
            if self.eugene.isChecked():
                eugene = []
                eugene_total_13 = 0
                eugene_total_14 = 0
                eugene_total_15 = 0
                eugene_total_16 = 0
                eugene_total_17 = 12.30000
                eugene_cal = []
                eugene_all = []
                
                eugene_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in eugene_eq:
                    i = df[(df.Competitor == 'Eugene Technology' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    eugene.append(ii)
                
                eugene_total_18 = round(sum(eugene),0)
                
                eugene_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                eugene_total = [eugene_total_13, eugene_total_14, eugene_total_15, eugene_total_16, eugene_total_17, eugene_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    eugene_cal.append(round((eugene_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(eugene_year, eugene_cal, color = 'green')
                ax.plot(eugene_year, eugene_cal, color ='green')
                
                ax.set_xticks(eugene_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % eugene_cal[num], xy = (eugene_year[num],eugene_cal[num]), textcoords='data', fontsize = 8)
                
                    yellow_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[yellow_patch,orange_patch,lightgrey_patch,blue_patch])
        
            if self.jusung.isChecked():
                jusung = []
                jusung_total_13 = 32.0
                jusung_total_14 = 51.0
                jusung_total_15 = 63.1
                jusung_total_16 = 86.900
                jusung_total_17 = 97.80
                jusung_all = []
                jusung_cal = []
                
                jusung_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in jusung_eq:
                    i = df[(df.Competitor == 'Jusung Engineering' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    jusung.append(ii)
                
                jusung_total_18 = round(sum(jusung),0)
                
                jusung_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                jusung_total = [jusung_total_13, jusung_total_14, jusung_total_15, jusung_total_16, jusung_total_17, jusung_total_18]

                
                for num in [0,1,2,3,4,5]:
                    jusung_cal.append(round((jusung_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(jusung_year,jusung_cal, color = 'aqua')
                ax.plot(jusung_year, jusung_cal, color ='aqua')
                
                ax.set_xticks(jusung_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' %jusung_cal[num], xy = (jusung_year[num],jusung_cal[num]), textcoords='data', fontsize = 8)
                    
                    aqua_patch = mpatches.Patch(color='aqua',label='Jusung')
                    yellow_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[aqua_patch,yellow_patch,orange_patch,lightgrey_patch,blue_patch])
        
            if self.lam.isChecked():
                lam = []
                lam_total_13 = 0.0
                lam_total_14 = 45.5700295
                lam_total_15 = 95.4
                lam_total_16 = 63.5875
                lam_total_17 = 214.8999
                lam_all = []
                lam_cal = []
                
                lam_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in lam_eq:
                    i = df[(df.Competitor == 'Lam Research' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    lam.append(ii)
                
                lam_total_18 = round(sum(lam),0)
                
                lam_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                lam_total = [lam_total_13,lam_total_14, lam_total_15, lam_total_16, lam_total_17, lam_total_18]
            
                
                for num in [0,1,2,3,4,5]:
                    lam_cal.append(round((lam_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(lam_year, lam_cal, color = 'red')
                ax.plot(lam_year, lam_cal, color ='red')
                
                ax.set_xticks(lam_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % lam_cal[num], xy = (lam_year[num],lam_cal[num]), textcoords='data', fontsize = 8)
                    
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    aqua_patch = mpatches.Patch(color='aqua',label='Jusung')
                    yellow_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[red_patch,aqua_patch,yellow_patch,orange_patch,lightgrey_patch,blue_patch])
    
            if self.ultratech.isChecked():
                ultra = []
                ultra_total_13 = 7.225
                ultra_total_14 = 5.155
                ultra_total_15 = 8.75
                ultra_total_16 = 9.1300
                ultra_total_17 = 7.4555
                ultra_all = []
                ultra_cal = []
                
                ultra_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in ultra_eq:
                    i = df[(df.Competitor == 'Ultratech' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    ultra.append(ii)
                
                ultra_total_18 = round(sum(ultra),0)
                
                ultra_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                ultra_total = [ultra_total_13, ultra_total_14, ultra_total_15, ultra_total_16, ultra_total_17, ultra_total_18]
            
                
                for num in [0,1,2,3,4,5]:
                    ultra_cal.append(round((ultra_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(ultra_year, ultra_cal, color = 'violet')
                ax.plot(ultra_year, ultra_cal, color ='violet')
                
                ax.set_xticks(ultra_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % ultra_cal[num], xy = (ultra_year[num],ultra_cal[num]), textcoords='data', fontsize = 8)
                    
                    violet_patch = mpatches.Patch(color='violet',label='Ultratech')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    aqua_patch = mpatches.Patch(color='aqua',label='Jusung')
                    yellow_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[violet_patch,red_patch,aqua_patch,yellow_patch,orange_patch,lightgrey_patch,blue_patch])
                    
            
            if self.veeco.isChecked():
                veeco = []
                veeco_total_13 = 0.0
                veeco_total_14 = 0.0
                veeco_total_15 = 0.0
                veeco_total_16 = 0.0
                veeco_total_17 = 4.115830898
                veeco_all = []
                veeco_cal = []
                
                veeco_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in veeco_eq:
                    i = df[(df.Competitor == 'Veeco' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    veeco.append(ii)
                
                veeco_total_18 = round(sum(veeco),0)
                
                veeco_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                veeco_total = [veeco_total_13, veeco_total_14, veeco_total_15, veeco_total_16, veeco_total_17, veeco_total_18]

                
                for num in [0,1,2,3,4,5]:
                    veeco_cal.append(round((veeco_total[num]/aldp_all[num])*100, 2))
            
                ax.scatter(veeco_year, veeco_cal, color = 'salmon')
                ax.plot(veeco_year, veeco_cal, color ='salmon')
                
                ax.set_xticks(veeco_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % veeco_cal[num], xy = (veeco_year[num],veeco_cal[num]), textcoords='data', fontsize = 8)
                    
                    salmon_patch = mpatches.Patch(color = 'salmon', label = 'Veeco')
                    violet_patch = mpatches.Patch(color='violet',label='Ultratech')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    aqua_patch = mpatches.Patch(color='aqua',label='Jusung')
                    yellow_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[salmon_patch,violet_patch,red_patch,aqua_patch,yellow_patch,orange_patch,lightgrey_patch,blue_patch])
        
            if self.wonik.isChecked():
                wonik = []
                wonik_total_13 = 39.4
                wonik_total_14 = 45.3
                wonik_total_15 = 47.7509
                wonik_total_16 = 45.60211
                wonik_total_17 = 123.59999
                wonik_all = []
                wonik_cal = []
                
                wonik_eq = ['Atomic Layer Deposition Platforms' ]
                
                for eq in wonik_eq:
                    i = df[(df.Competitor == 'Wonik IPS' )&(df.Yearly>0) & (df.CY == 'CY18') & (df['Level 2'] ==  eq)]
                    ii = float(i['Yearly'])
                    wonik.append(ii)
                
                wonik_total_18 = round(sum(wonik),0)
                
                wonik_year = ['CY13','CY14','CY15','CY16','CY17', 'CY18']
                
                wonik_total = [wonik_total_13, wonik_total_14, wonik_total_15, wonik_total_16, wonik_total_17, wonik_total_18]
                
                
                for num in [0,1,2,3,4,5]:
                    wonik_cal.append(round((wonik_total[num]/aldp_all[num])*100, 2))
                    
                ax.scatter(wonik_year, wonik_cal, color = 'goldenrod')
                ax.plot(wonik_year, wonik_cal, color ='goldenrod')
                
                ax.set_xticks(wonik_year)
            
                ax.set_xlabel('Year')
                ax.set_title("Segment competitor by shares")
                ax.set_ylabel('Shares(in percent)')
                ax.yaxis.set_major_formatter(PercentFormatter())
                
                for num in [0,1,2,3,4,5]:                               # <--
                    ax.annotate('(%s)' % wonik_cal[num], xy = (wonik_year[num],wonik_cal[num]), textcoords='data', fontsize = 8)
                    golden_patch = mpatches.Patch(color='goldenrod',label='Wonik')
                    salmon_patch = mpatches.Patch(color = 'salmon', label = 'Veeco')
                    violet_patch = mpatches.Patch(color='violet',label='Ultratech')
                    red_patch = mpatches.Patch(color='red',label='Lam')
                    aqua_patch = mpatches.Patch(color='aqua',label='Jusung')
                    yellow_patch = mpatches.Patch(color='green',label='Eugene')
                    orange_patch = mpatches.Patch(color='orange',label='ASM')
                    lightgrey_patch = mpatches.Patch(color='lightgrey',label='Aixtron')
                    blue_patch = mpatches.Patch(color='blue',label='AMAT')
                    ax.legend(handles=[golden_patch,salmon_patch,violet_patch,red_patch,aqua_patch,yellow_patch,orange_patch,lightgrey_patch,blue_patch])
        # refresh canvas
        self.canvas.draw()
 
class Third(QDialog):
    def __init__(self, parent = None):
        super(Third, self).__init__(parent)
        self.createFormGroupBox()
        
    def createFormGroupBox(self):  
        
        font = QFont()
        font.setPointSize(16)
        
        year = QHBoxLayout()
        
        load_previous = QHBoxLayout()
        
        update = QHBoxLayout()
        nlpcvd_value = QHBoxLayout()
        
        pcvd_value = QHBoxLayout()
        
        aldp_value = QHBoxLayout()
        
        tcvd_value = QHBoxLayout()
        
        buttonLayout1 = QVBoxLayout()  
        buttons = QHBoxLayout()
        
        nlpcvd = QLabel("NLPCVD:")
        pcvd = QLabel("PCVD:")
        aldp = QLabel("ALDP:")
        tcvd = QLabel("TCVD:")
        
        a = QLabel("Lock")
        aa = QLabel("Lock")
        aaa = QLabel("Lock")
        aaaa = QLabel("Lock")
        
        not_resize = a.sizePolicy();
        not_resize.setRetainSizeWhenHidden(True);
        a.setSizePolicy(not_resize);
        a.setVisible(True)
                     
        self.year = QPushButton("Select the year:")
        self.year.clicked.connect(self.getItem)
        
        self.load = QPushButton("Load previous values")
        
        self.update = QPushButton("Update")
        
        self.namebox = QCheckBox()
        self.namebox1 = QCheckBox()
        self.namebox2 = QCheckBox()
        self.namebox3 = QCheckBox()

        self.nameLine = QLineEdit()
        
        self.nameLine1 = QLineEdit()
        self.nameLine2 = QLineEdit()
        self.nameLine3 = QLineEdit()
        self.nameLine4 = QLineEdit()
        
        self.verified1 = QLineEdit()
        self.verified2 = QLineEdit()
        self.verified3 = QLineEdit()
        self.verified4 = QLineEdit()
        
                 
        self.cancel = QPushButton("Finish")
        self.cancel.clicked.connect(self.submitContact)
        
        self.namebox.setChecked(False)

        self.namebox1.setChecked(False)

        self.namebox2.setChecked(False)

        self.namebox3.setChecked(False)
        
        self.setFont(font)
                
        year.addWidget(self.year)
        year.addWidget(self.nameLine)
        buttonLayout1.addLayout(year)
        
        load_previous.addWidget(self.load)
        self.load.clicked.connect(self.nlpcvd)
        self.load.clicked.connect(self.pcvd)
        self.load.clicked.connect(self.aldp)
        self.load.clicked.connect(self.tcvd)
        
        load_previous.addWidget(self.update)
        self.update.clicked.connect(self.checked)
        self.update.clicked.connect(self.new_nlpcvd)
        self.update.clicked.connect(self.new_pcvd)
        self.update.clicked.connect(self.new_aldp)
        self.update.clicked.connect(self.new_tcvd)
        buttonLayout1.addLayout(load_previous)
        
        nlpcvd_value.addWidget(nlpcvd)
        nlpcvd_value.addWidget(self.nameLine1)
        nlpcvd_value.addWidget(self.verified1)
        nlpcvd_value.addWidget(a)
        nlpcvd_value.addWidget(self.namebox)
        
        buttonLayout1.addLayout(nlpcvd_value)
        
        pcvd_value.addWidget(pcvd)
        pcvd_value.addWidget(self.nameLine2)
        pcvd_value.addWidget(self.verified2)
        pcvd_value.addWidget(aa)
        pcvd_value.addWidget(self.namebox1)
        
        buttonLayout1.addLayout(pcvd_value)
        
        aldp_value.addWidget(aldp)
        aldp_value.addWidget(self.nameLine3)
        aldp_value.addWidget(self.verified3)
        aldp_value.addWidget(aaa)
        aldp_value.addWidget(self.namebox2)
        
        buttonLayout1.addLayout(aldp_value)
        
        tcvd_value.addWidget(tcvd)
        tcvd_value.addWidget(self.nameLine4)
        tcvd_value.addWidget(self.verified4)
        tcvd_value.addWidget(aaaa)
        tcvd_value.addWidget(self.namebox3)
        
        buttonLayout1.addLayout(tcvd_value)
        
        buttons.addWidget(self.cancel)
        buttonLayout1.addLayout(buttons)   
        
        mainLayout = QGridLayout()
        # mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addLayout(buttonLayout1, 0, 1)
 
        self.setLayout(mainLayout)
        self.setLayout(buttonLayout1)
        self.setWindowTitle("Hello Qt")
        
    def checked(self):
        q1 = (self.nameLine.text())
        q2 = (self.nameLine1.text())
        q3 = (self.nameLine2.text())
        q4 = (self.nameLine3.text())
        q5 = (self.nameLine4.text())
        if q1 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D1'] = q2
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E1'] = q3
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F1'] = q4
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G1'] = q5
                
                wb1.save(filepath)
    
    def getItem(self):
        items1 = ("2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025")
		
        items, ok = QInputDialog.getItem(self, "select input dialog", 
         "list of competitors", items1, 0, False)
			
        if ok and items:
            self.nameLine.setText(items)
            print(items)
    
    def nlpcvd(self):
        name = (self.nameLine.text())
        
        if name == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            
            q2 = ws1['D1'].value
            
            self.nameLine1.setText(q2)
    
    def pcvd(self):
        name = (self.nameLine.text())
        
        if name == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            
            q2 = ws1['E1'].value
            
            self.nameLine2.setText(q2)
                
    def aldp(self):
        name = (self.nameLine.text())
        
        if name == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            
            q2 = ws1['F1'].value
            
            self.nameLine3.setText(q2)
                
    def tcvd(self):
        name = (self.nameLine.text())
        
        if name == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            
            q2 = ws1['G1'].value
            
            self.nameLine4.setText(q2)

    def new_nlpcvd(self):
        year = (self.nameLine.text())
        nlpcvd = self.nameLine1.text()
        nlpcvd = float(nlpcvd)
        
        if year == '2018': 
            print("yes")
            if nlpcvd != "":
                filepath = 'Competitor-Flat-File.xlsx'
                wb = load_workbook(filepath)
                ws = wb['Revised']
                
                ws['F489'] = nlpcvd   
                
                wb.save(filepath)
                
                self.verified1.setText("%s" % nlpcvd)
                    
                
    def new_pcvd(self):
        year = (self.nameLine.text())
        nlpcvd = self.nameLine1.text()
        nlpcvd = float(nlpcvd)
        pcvd = self.nameLine2.text()
        pcvd = float(pcvd)
            
        if year == '2018':
            if pcvd != "":
                filepath = 'Competitor-Flat-File.xlsx'
                wb = load_workbook(filepath)
                ws = wb['Revised']
                
                ws['F490'] = pcvd  
                
                wb.save(filepath)
                
                self.verified1.setText("%s" % nlpcvd)
                self.verified2.setText("%s" % pcvd)

                
    def new_aldp(self):
        year = (self.nameLine.text())
        nlpcvd = self.nameLine1.text()
        nlpcvd = float(nlpcvd)
        pcvd = self.nameLine2.text()
        pcvd = float(pcvd)
        aldp = self.nameLine3.text()
        aldp = float(aldp)
        
        if year == '2018':
            if aldp != "":
                filepath = 'Competitor-Flat-File.xlsx'
                wb = load_workbook(filepath)
                ws = wb['Revised']
                
                ws['F487'] = aldp
                
                wb.save(filepath)
                
                self.verified1.setText("%s" % nlpcvd)
                self.verified2.setText("%s" % pcvd)
                self.verified3.setText("%s" % aldp)
                
    def new_tcvd(self):
        year = (self.nameLine.text())
        nlpcvd = self.nameLine1.text()
        nlpcvd = float(nlpcvd)
        pcvd = self.nameLine2.text()
        pcvd = float(pcvd)
        aldp = self.nameLine3.text()
        aldp = float(aldp)
        tcvd = self.nameLine4.text()
        tcvd = float(tcvd)
        
        if year == '2018':
            if tcvd != "":
                filepath = 'Competitor-Flat-File.xlsx'
                wb = load_workbook(filepath)
                ws = wb['Revised']
                
                ws['F484'] = tcvd
                
                wb.save(filepath)
                
                self.verified1.setText("%s" % nlpcvd)
                self.verified2.setText("%s" % pcvd)
                self.verified3.setText("%s" % aldp)
                self.verified4.setText("%s" % tcvd)
    
    def submitContact(self):
        self.close()
        
class Second(QDialog):
    def __init__(self, parent = None):
        super(Second, self).__init__(parent)
        self.createFormGroupBox()
        
    def createFormGroupBox(self): 
        
        font = QFont()
        font.setPointSize(16)
        
        year = QHBoxLayout()
        comp = QHBoxLayout()
        ratio1 = QHBoxLayout()
                
        add_quarter = QHBoxLayout()
        
        hboxa = QHBoxLayout()
        hboxa1 = QHBoxLayout()
        hboxa2 = QHBoxLayout()
        hboxa3 = QHBoxLayout()
                
        buttons = QHBoxLayout()
        
        buttonLayout1 = QVBoxLayout()   
        
        ratio = QLabel("Ratio for Q1 & Q2:")
        
        nameLabel = QLabel("Quarter 1:")
        nameLabel1 = QLabel("Quarter 2:")
        nameLabel2 = QLabel("Quarter 3:")
        nameLabel3 = QLabel("Quarter 4:")
        
        self.pre_quarter = QPushButton("Load quarters")
        
        self.add_quarter = QPushButton("Update quarters")
        
        a = QLabel("Lock")
        aa = QLabel("Lock")
        aaa = QLabel("Lock")
        aaaa = QLabel("Lock")
        aaaaa = QLabel("Lock")
                     
        # size policy
        not_resize = a.sizePolicy();
        not_resize.setRetainSizeWhenHidden(True);
        a.setSizePolicy(not_resize);
        a.setVisible(True)
        
        self.competitor = QPushButton("Select the competitor:")
        self.competitor.clicked.connect(self.getItem)
        
        self.year = QLabel("Year selected:")
        
        self.setFont(font)
        
        self.ratio = QLineEdit()
        
        self.namebox = QCheckBox()
        self.namebox1 = QCheckBox()
        self.namebox2 = QCheckBox()
        self.namebox3 = QCheckBox()
        self.ratio_box = QCheckBox()
        self.nameLine = QLineEdit()
        self.nameLine1 = QLineEdit()
        self.nameLine2 = QLineEdit()
        self.nameLine3 = QLineEdit()
        self.nameLine4 = QLineEdit()
        self.nameLine5 = QLineEdit()
        self.verified1 = QLineEdit()
        self.verified2 = QLineEdit()
        self.verified3 = QLineEdit()
        self.verified4 = QLineEdit()
        
        self.ratio_graph = QPushButton("Ratio graph")
        self.ratio_graph.clicked.connect(self.graph_ratio)
        
        self.submitButton = QPushButton("Finish")
        self.submitButton.clicked.connect(self.submitContact)
        
        comp.addWidget(self.competitor)
        comp.addWidget(self.nameLine4)
        buttonLayout1.addLayout(comp)
        
        year.addWidget(self.year)
        year.addWidget(self.nameLine5)
        buttonLayout1.addLayout(year)

        self.namebox.setChecked(False)

        self.namebox1.setChecked(False)

        self.namebox2.setChecked(False)

        self.namebox3.setChecked(False)
        
        #self.nameLine.setReadOnly(self.namebox.checkState()==Qt.Checked)
        #self.namebox.stateChanged.connect(lambda state: self.nameLine.setReadOnly(state==Qt.Checked))
        #self.namebox.setChecked(True)
        
        #self.nameLine1.setReadOnly(self.namebox1.checkState()==Qt.Checked)
        #self.namebox1.stateChanged.connect(lambda state: self.nameLine1.setReadOnly(state==Qt.Checked))
        #self.namebox1.setChecked(True)
        
        #self.ratio.setReadOnly(self.namebox.checkState()==Qt.Checked)
        #self.ratio_box.stateChanged.connect(lambda state: self.ratio.setReadOnly(state==Qt.Checked))
        #self.ratio_box.setChecked(True)
        
        add_quarter.addWidget(self.pre_quarter)
        self.pre_quarter.clicked.connect(self.prequarter1)
        self.pre_quarter.clicked.connect(self.prequarter2)
        self.pre_quarter.clicked.connect(self.prequarter3)
        self.pre_quarter.clicked.connect(self.prequarter4)
        self.pre_quarter.clicked.connect(self.set_ratio)
        
        add_quarter.addWidget(self.add_quarter)  
        buttonLayout1.addLayout(add_quarter)
        self.add_quarter.clicked.connect(self.quarter1)
        self.add_quarter.clicked.connect(self.quarter2)
        self.add_quarter.clicked.connect(self.quarter3)
        self.add_quarter.clicked.connect(self.quarter4)
        self.add_quarter.clicked.connect(self.checked)
        
        hboxa.addWidget(nameLabel)
        hboxa.addWidget(self.nameLine)
        hboxa.addWidget(self.verified1)
        hboxa.addWidget(a)
        hboxa.addWidget(self.namebox)
        buttonLayout1.addLayout(hboxa)
        
        hboxa1.addWidget(nameLabel1)
        hboxa1.addWidget(self.nameLine1)
        hboxa1.addWidget(self.verified2)
        hboxa1.addWidget(aa)
        hboxa1.addWidget(self.namebox1)
        buttonLayout1.addLayout(hboxa1)
        
        hboxa2.addWidget(nameLabel2)
        hboxa2.addWidget(self.nameLine2)
        hboxa2.addWidget(self.verified3)
        hboxa2.addWidget(aaa)
        hboxa2.addWidget(self.namebox2)
        buttonLayout1.addLayout(hboxa2)
        
        hboxa3.addWidget(nameLabel3)
        hboxa3.addWidget(self.nameLine3)
        hboxa3.addWidget(self.verified4)
        hboxa3.addWidget(aaaa)
        hboxa3.addWidget(self.namebox3)
        buttonLayout1.addLayout(hboxa3)
        
        ratio1.addWidget(ratio)
        ratio1.addWidget(self.ratio)
        ratio1.addWidget(aaaaa)
        ratio1.addWidget(self.ratio_box)
        buttonLayout1.addLayout(ratio1)
        
        buttons.addWidget(self.ratio_graph)
        buttons.addWidget(self.submitButton)
        buttonLayout1.addLayout(buttons)
        
        mainLayout = QGridLayout()
        # mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addLayout(buttonLayout1, 0, 1)
 
        self.setLayout(mainLayout)
        self.setLayout(buttonLayout1)
        self.setWindowTitle("Hello Qt")
    
    def checked(self):
        q1 = (self.nameLine.text())
        q2 = (self.nameLine1.text())
        q3 = (self.nameLine2.text())
        q4 = (self.nameLine3.text())
        q5 = (self.ratio.text())
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        if name1 == 'Aixtron' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D2'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E2'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F2'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G2'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H2'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'ASM International' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D3'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E3'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F3'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G3'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H3'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Eugene Technology' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D4'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E4'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F4'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G4'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H4'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D5'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E5'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F5'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G5'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H5'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Jusung Engineering' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D6'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E6'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F6'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G6'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H6'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Lam Research' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D7'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E7'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F7'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G7'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H7'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Orbotech' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D8'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E8'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F8'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G8'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H8'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Tes' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D9'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E9'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F9'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G9'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H9'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Tokyo Electron' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D10'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E10'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F10'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G10'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H10'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Ultratech' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D11'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E11'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F11'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G11'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H11'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Veeco' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D12'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E12'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F12'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G12'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H12'] = q5
                
                wb1.save(filepath)
                
        if name1 == 'Wonik IPS' and name2 == '2018':
            if self.namebox.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['D13'] = q1
                
                wb1.save(filepath)
            if self.namebox1.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['E13'] = q2
                
                wb1.save(filepath)
            if self.namebox2.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['F13'] = q3
                
                wb1.save(filepath)
            if self.namebox3.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['G13'] = q4
                
                wb1.save(filepath)
            
            if self.ratio_box.isChecked():
                filepath = 'Competitor-Flat-File.xlsx'
                wb1 = load_workbook(filepath)
                ws1 = wb1['QoQ']
                
                ws1['H13'] = q5
                
                wb1.save(filepath)
    
    def getItem(self):       
        items = ("Aixtron", "ASM International", "Eugene Technology", "Hitachi Kokusai Electric", "Jusung Engineering", "Lam Research", "Orbotech", "Tes", "Tokyo Electron", "Ultratech", "Veeco", "Wonik IPS")
		
        item, ok = QInputDialog.getItem(self, "select input dialog", 
         "list of competitors", items, 0, False)
			
        if ok and item:
            self.nameLine4.setText(item)
            competitor.append(item)
            print(item)
            
        items1 = ("2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025")
		
        items, ok = QInputDialog.getItem(self, "select input dialog", 
         "list of years", items1, 0, False)
			
        if ok and items:
            self.nameLine5.setText(items)
            year.append(items)
            print(items) 
    
    def prequarter1(self):
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        
        if name1 == 'Aixtron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D2'].value
            
            self.nameLine.setText(q1)
            
                
        if name1 == 'ASM International' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D3'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Eugene Technology' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D4'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D5'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Jusung Engineering' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D6'].value
            
            self.nameLine.setText(q1)
        
        if name1 == 'Lam Research' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D7'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Orbotech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D8'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Tes' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D9'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Tokyo Electron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D10'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Ultratech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D11'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Veeco' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D12'].value
            
            self.nameLine.setText(q1)
                
        if name1 == 'Wonik IPS' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['D13'].value
            
            self.nameLine.setText(q1)
            
    def prequarter2(self):
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        
        if name1 == 'Aixtron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E2'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'ASM International' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E3'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Eugene Technology' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E4'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E5'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Jusung Engineering' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E6'].value
            
            self.nameLine1.setText(q1)
        
        if name1 == 'Lam Research' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E7'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Orbotech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E8'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Tes' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E9'].value
            
            self.nameLine1.setText(q1)
            
        if name1 == 'Tokyo Electron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E10'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Ultratech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E11'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Veeco' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E12'].value
            
            self.nameLine1.setText(q1)
                
        if name1 == 'Wonik IPS' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['E13'].value
            
            self.nameLine1.setText(q1)
            
    def prequarter3(self):
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        
        if name1 == 'Aixtron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F2'].value
            
            self.nameLine2.setText(q1)
        
        if name1 == 'ASM International' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F3'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Eugene Technology' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F4'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F5'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Jusung Engineering' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F6'].value
            
            self.nameLine2.setText(q1)
        
        if name1 == 'Lam Research' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F7'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Orbotech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F8'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Tes' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F9'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Tokyo Electron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F10'].value
            
            self.nameLine2.setText(q1)
        
        if name1 == 'Ultratech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F11'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Veeco' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F12'].value
            
            self.nameLine2.setText(q1)
                
        if name1 == 'Wonik IPS' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['F13'].value
            
            self.nameLine2.setText(q1)
    
    def prequarter4(self):
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        
        if name1 == 'Aixtron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G2'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'ASM International' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G3'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Eugene Technology' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G4'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G5'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Jusung Engineering' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G6'].value
            
            self.nameLine3.setText(q1)
        
        if name1 == 'Lam Research' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G7'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Orbotech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G8'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Tes' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G9'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Tokyo Electron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G10'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Ultratech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G11'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Veeco' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G12'].value
            
            self.nameLine3.setText(q1)
                
        if name1 == 'Wonik IPS' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['G13'].value
            
            self.nameLine3.setText(q1)

    def set_ratio(self):
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        
        if name1 == 'Aixtron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H2'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'ASM International' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H3'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Eugene Technology' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H4'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H5'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Jusung Engineering' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H6'].value
            
            self.ratio.setText(q1)
        
        if name1 == 'Lam Research' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H7'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Orbotech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H8'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Tes' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H9'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Tokyo Electron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H10'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Ultratech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H11'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Veeco' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H12'].value
            
            self.ratio.setText(q1)
                
        if name1 == 'Wonik IPS' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb1 = load_workbook(filepath)
            ws1 = wb1['QoQ']
            q1 = ws1['H13'].value
            
            self.ratio.setText(q1)
        
    def quarter1(self):
        pre_quarter_1 = self.nameLine.text()
        quarter_1 = float(pre_quarter_1)
        pre_ratio = self.ratio.text()
        ratio = float(pre_ratio)
        
        for item, items in zip(competitor, year):
            if item == 'Aixtron' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = round(float(quarter_1), 3)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F1021'] = total   
                        
                    tab = ['F1031', 'F1032', 'F1035', 'F1037', 'F1038', 'F1039', 'F1041']
                                        
                    ratio = [(ws['F946'].value/ws['F936'].value)*ws['F1021'].value, (ws['F947'].value/ws['F936'].value)*ws['F1021'].value, (ws['F950'].value/ws['F936'].value)*ws['F1021'].value, (ws['F952'].value/ws['F936'].value)*ws['F1021'].value, (ws['F953'].value/ws['F936'].value)*ws['F1021'].value, (ws['F954'].value/ws['F936'].value)*ws['F1021'].value, (ws['F956'].value/ws['F936'].value)*ws['F1021'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                                        
        
        for item, items in zip(competitor, year):
            if item == 'ASM International' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.22))
                    
                    ws['F1573'] = total  
                    
                    tab = ['F1583', 'F1585', 'F1587', 'F1588', 'F1589', 'F1590', 'F1591', 'F1594', 'F1618', 'F1622']
                                        
                    ratio = [(ws['F1498'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1500'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1502'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1503'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1504'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1505'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1506'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1509'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1533'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1537'].value/ws['F1488'].value)*ws['F1573'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Eugene Technology' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.84))
                    
                    ws['F2125'] = total   
                    
                    tab = ['F2135', 'F2137', 'F2139', 'F2141', 'F2142', 'F2143', 'F2145', 'F2170', 'F2173']
                                        
                    ratio = [(ws['F2050'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2052'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2054'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2056'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2057'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2058'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2060'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2085'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2088'].value/ws['F2040'].value)*ws['F2125'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Hitachi Kokusai Electric' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F2677'] = total 
                    
                    tab = ['F2687', 'F2691', 'F2692', 'F2704', 'F2713', 'F2722', 'F2723', 'F2725', 'F2726']
                                        
                    ratio = [(ws['F2602'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2606'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2607'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2619'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2628'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2637'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2638'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2640'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2641'].value/ws['F2592'].value)*ws['F2677'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Jusung Engineering' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.40))
                    
                    ws['F3229'] = total   
                    
                    tab = ['F3239', 'F3243', 'F3245', 'F3246', 'F3247']
                                        
                    ratio = [(ws['F3154'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3158'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3160'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3161'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3162'].value/ws['F3144'].value)*ws['F3229'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
        
        for item, items in zip(competitor, year):
            if item == 'Lam Research' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F3781'] = total   
                    
                    tab = ['F3791', 'F3795', 'F3797', 'F3798', 'F3799', 'F3801', 'F3802', 'F3804', 'F3808', 'F3809', 'F3811', 'F3812', 'F3817', 'F3818', 'F3819', 'F3820', 'F3821']
                                        
                    ratio = [(ws['F3706'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3710'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3712'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3713'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3714'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3716'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3717'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3719'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3723'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3724'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3726'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3727'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3732'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3733'].value/ws['F3696'].value)*ws['F3781'].value,(ws['F3734'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3735'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3736'].value/ws['F3696'].value)*ws['F3781'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Orbotech' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F4307'] = total
                    
                    tab = ['F4317', 'F4321', 'F4323', 'F4328', 'F4329', 'F4334', 'F4344', 'F4345', 'F4347', 'F4348', 'F4350', 'F4389', 'F4391']
                                        
                    ratio = [(ws['F4232'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4236'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4238'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4243'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4244'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4249'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4259'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4260'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4262'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4263'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4265'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4304'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4306'].value/ws['F4222'].value)*ws['F4307'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Tes' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.92))
                    
                    ws['F4859'] = total  
                    
                    tab = ['F4869', 'F4873', 'F4875', 'F4879', 'F4880', 'F4886', 'F4892', 'F4894']
                                        
                    ratio = [(ws['F4784'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4788'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4790'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4794'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4795'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4801'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4807'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4809'].value/ws['F4774'].value)*ws['F4859'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Tokyo Electron' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.17))
                    
                    ws['F5411'] = total   
                    
                    tab = ['F5420', 'F5421', 'F5425', 'F5426', 'F5427', 'F5428', 'F5430', 'F5431', 'F5432', 'F5433', 'F5434', 'F5438', 'F5440', 'F5441', 'F5442', 'F5443', 'F5444', 'F5445', 'F5446', 'F5448', 'F5449', 'F5450', 'F5451', 'F5456', 'F5459', 'F5460', 'F5493', 'F5495']
                                        
                    ratio = [(ws['F5335'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5336'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5340'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5341'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5342'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5343'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5345'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5346'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5347'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5348'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5349'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5353'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5355'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5356'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5357'].value/ws['F5326'].value)*ws['F5411'].value,(ws['F5358'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5359'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5360'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5361'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5363'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5364'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5365'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5366'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5371'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5374'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5375'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5408'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5410'].value/ws['F5326'].value)*ws['F5411'].value] 
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Ultratech' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.17))
                    
                    ws['F5963'] = total   
                    
                    tab = ['F5964', 'F5966', 'F5970', 'F5973', 'F5977', 'F5979', 'F5980', 'F5981', 'F6008', 'F6009', 'F6018', 'F6024', 'F6026']
                                        
                    ratio = [(ws['F5879'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5881'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5886'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5888'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5892'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5894'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5895'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5896'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5923'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5924'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5933'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5939'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5941'].value/ws['F5878'].value)*ws['F5963'].value] 
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Veeco' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.17))
                    
                    ws['F6515'] = total   
                    
                    tab = ['F6516', 'F6518', 'F6523', 'F6525', 'F6526', 'F6529', 'F6531', 'F6532', 'F6533', 'F6539', 'F6540', 'F6560', 'F6561', 'F6570', 'F6576', 'F6578']
                                        
                    ratio = [(ws['F6431'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6433'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6438'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6440'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6441'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6444'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6446'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6447'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6448'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6454'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6455'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6475'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6476'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6485'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6491'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6493'].value/ws['F6430'].value)*ws['F6515'].value ] 
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Wonik IPS' and items == '2018': 
                if quarter_1 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_1)
                    quart3 = quarter_1 * ratio_cal
                    quart4 = quarter_1 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.34))
                    
                    ws['F7067'] = total
                    
                    tab = ['F7077', 'F7081', 'F7083', 'F7084', 'F7085', 'F7087', 'F7088']
                                        
                    ratio = [(ws['F6992'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6996'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6998'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6999'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7000'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7002'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7003'].value/ws['F6982'].value)*ws['F7067'].value ] 
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
    def quarter2(self):
        pre_quarter_1 = self.nameLine.text()
        quarter_1 = float(pre_quarter_1)
        pre_quarter_2 = self.nameLine1.text()
        quarter_2 = float(pre_quarter_2)
        pre_ratio = self.ratio.text()
        ratio = float(pre_ratio)
        
        for item, items in zip(competitor, year):
            if item == 'Aixtron' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F1021'] = total
                    
                    tab = ['F1031', 'F1032', 'F1035', 'F1037', 'F1038', 'F1039', 'F1041']
                                        
                    ratio = [(ws['F946'].value/ws['F936'].value)*ws['F1021'].value, (ws['F947'].value/ws['F936'].value)*ws['F1021'].value, (ws['F950'].value/ws['F936'].value)*ws['F1021'].value, (ws['F952'].value/ws['F936'].value)*ws['F1021'].value, (ws['F953'].value/ws['F936'].value)*ws['F1021'].value, (ws['F954'].value/ws['F936'].value)*ws['F1021'].value, (ws['F956'].value/ws['F936'].value)*ws['F1021'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'ASM International' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                                                
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.22))
                    
                    ws['F1573'] = total   
                    
                    tab = ['F1583', 'F1585', 'F1587', 'F1588', 'F1589', 'F1590', 'F1591', 'F1594', 'F1618', 'F1622']
                                        
                    ratio = [(ws['F1498'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1500'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1502'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1503'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1504'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1505'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1506'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1509'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1533'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1537'].value/ws['F1488'].value)*ws['F1573'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
        
        for item, items in zip(competitor, year):
            if item == 'Eugene Technology' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                     
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.84))
                    
                    ws['F2125'] = total   
                    
                    tab = ['F2135', 'F2137', 'F2139', 'F2141', 'F2142', 'F2143', 'F2145', 'F2170', 'F2173']
                                        
                    ratio = [(ws['F2050'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2052'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2054'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2056'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2057'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2058'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2060'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2085'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2088'].value/ws['F2040'].value)*ws['F2125'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                     
        for item, items in zip(competitor, year):
            if item == 'Hitachi Kokusai Electric' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F2677'] = total   
                    
                    tab = ['F2687', 'F2691', 'F2692', 'F2704', 'F2713', 'F2722', 'F2723', 'F2725', 'F2726']
                                        
                    ratio = [(ws['F2602'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2606'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2607'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2619'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2628'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2637'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2638'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2640'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2641'].value/ws['F2592'].value)*ws['F2677'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Jusung Engineering' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.40))
                    
                    ws['F3229'] = total  
                    
                    tab = ['F3239', 'F3243', 'F3245', 'F3246', 'F3247']
                                        
                    ratio = [(ws['F3154'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3158'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3160'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3161'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3162'].value/ws['F3144'].value)*ws['F3229'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)

        for item, items in zip(competitor, year):
            if item == 'Lam Research' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F3781'] = total   
                    
                    tab = ['F3791', 'F3795', 'F3797', 'F3798', 'F3799', 'F3801', 'F3802', 'F3804', 'F3808', 'F3809', 'F3811', 'F3812', 'F3817', 'F3818', 'F3819', 'F3820', 'F3821']
                                        
                    ratio = [(ws['F3706'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3710'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3712'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3713'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3714'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3716'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3717'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3719'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3723'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3724'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3726'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3727'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3732'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3733'].value/ws['F3696'].value)*ws['F3781'].value,(ws['F3734'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3735'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3736'].value/ws['F3696'].value)*ws['F3781'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Orbotech' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F4307'] = total 
                    
                    tab = ['F4317', 'F4321', 'F4323', 'F4328', 'F4329', 'F4334', 'F4344', 'F4345', 'F4347', 'F4348', 'F4350', 'F4389', 'F4391']
                                        
                    ratio = [(ws['F4232'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4236'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4238'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4243'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4244'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4249'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4259'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4260'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4262'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4263'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4265'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4304'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4306'].value/ws['F4222'].value)*ws['F4307'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Tes' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                             
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.92))
                    
                    ws['F4859'] = total   
                    
                    tab = ['F4869', 'F4873', 'F4875', 'F4879', 'F4880', 'F4886', 'F4892', 'F4894']
                                        
                    ratio = [(ws['F4784'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4788'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4790'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4794'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4795'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4801'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4807'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4809'].value/ws['F4774'].value)*ws['F4859'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Tokyo Electron' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = quarter_1 
                    quart2 = quarter_2
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.17))
                    
                    ws['F5411'] = total  
                    
                    tab = ['F5420', 'F5421', 'F5425', 'F5426', 'F5427', 'F5428', 'F5430', 'F5431', 'F5432', 'F5433', 'F5434', 'F5438', 'F5440', 'F5441', 'F5442', 'F5443', 'F5444', 'F5445', 'F5446', 'F5448', 'F5449', 'F5450', 'F5451', 'F5456', 'F5459', 'F5460', 'F5493', 'F5495']
                                        
                    ratio = [(ws['F5335'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5336'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5340'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5341'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5342'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5343'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5345'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5346'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5347'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5348'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5349'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5353'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5355'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5356'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5357'].value/ws['F5326'].value)*ws['F5411'].value,(ws['F5358'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5359'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5360'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5361'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5363'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5364'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5365'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5366'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5371'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5374'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5375'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5408'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5410'].value/ws['F5326'].value)*ws['F5411'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Ultratech' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                             
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F5963'] = total
                    
                    tab = ['F5964', 'F5966', 'F5970', 'F5973', 'F5977', 'F5979', 'F5980', 'F5981', 'F6008', 'F6009', 'F6018', 'F6024', 'F6026']
                                        
                    ratio = [(ws['F5879'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5881'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5886'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5888'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5892'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5894'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5895'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5896'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5923'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5924'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5933'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5939'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5941'].value/ws['F5878'].value)*ws['F5963'].value]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Veeco' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1)
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F6515'] = total 
                    
                    tab = ['F6516', 'F6518', 'F6523', 'F6525', 'F6526', 'F6529', 'F6531', 'F6532', 'F6533', 'F6539', 'F6540', 'F6560', 'F6561', 'F6570', 'F6576', 'F6578']
                                        
                    ratio = [(ws['F6431'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6433'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6438'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6440'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6441'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6444'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6446'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6447'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6448'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6454'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6455'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6475'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6476'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6485'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6491'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6493'].value/ws['F6430'].value)*ws['F6515'].value ]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Wonik IPS' and items == '2018': 
                if quarter_2 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_3 = (quarter_1+quarter_2)/2
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = quarter_3 * ratio_cal
                    quart4 = quarter_3 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.34))
                    
                    ws['F7067'] = total   
                    
                    tab = ['F7077', 'F7081', 'F7083', 'F7084', 'F7085', 'F7087', 'F7088']
                                        
                    ratio = [(ws['F6992'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6996'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6998'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6999'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7000'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7002'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7003'].value/ws['F6982'].value)*ws['F7067'].value ]
                                       
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
    
    def quarter3(self):
        pre_quarter_1 = self.nameLine.text()
        quarter_1 = float(pre_quarter_1)
        pre_quarter_2 = self.nameLine1.text()
        quarter_2 = float(pre_quarter_2)
        pre_quarter_3 = self.nameLine2.text()
        quarter_3 = float(pre_quarter_3)
        pre_ratio = self.ratio.text()
        ratio = float(pre_ratio)
        
        for item, items in zip(competitor, year):
            if item == 'Aixtron' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F1021'] = total   
                    
                    tab = ['F1031', 'F1032', 'F1035', 'F1037', 'F1038', 'F1039', 'F1041']
                                        
                    ratio = [(ws['F946'].value/ws['F936'].value)*ws['F1021'].value, (ws['F947'].value/ws['F936'].value)*ws['F1021'].value, (ws['F950'].value/ws['F936'].value)*ws['F1021'].value, (ws['F952'].value/ws['F936'].value)*ws['F1021'].value, (ws['F953'].value/ws['F936'].value)*ws['F1021'].value, (ws['F954'].value/ws['F936'].value)*ws['F1021'].value, (ws['F956'].value/ws['F936'].value)*ws['F1021'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                       
        for item, items in zip(competitor, year):
            if item == 'ASM International' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              

                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.22))
                    
                    ws['F1573'] = total 
                    
                    tab = ['F1583', 'F1585', 'F1587', 'F1588', 'F1589', 'F1590', 'F1591', 'F1594', 'F1618', 'F1622']
                                        
                    ratio = [(ws['F1498'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1500'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1502'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1503'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1504'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1505'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1506'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1509'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1533'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1537'].value/ws['F1488'].value)*ws['F1573'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Eugene Technology' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                                                
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.84))
                    
                    ws['F2125'] = total   
                    
                    tab = ['F2135', 'F2137', 'F2139', 'F2141', 'F2142', 'F2143', 'F2145', 'F2170', 'F2173']
                                        
                    ratio = [(ws['F2050'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2052'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2054'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2056'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2057'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2058'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2060'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2085'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2088'].value/ws['F2040'].value)*ws['F2125'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Hitachi Kokusai Electric' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F2677'] = total   
                    
                    tab = ['F2687', 'F2691', 'F2692', 'F2704', 'F2713', 'F2722', 'F2723', 'F2725', 'F2726']
                                        
                    ratio = [(ws['F2602'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2606'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2607'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2619'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2628'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2637'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2638'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2640'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2641'].value/ws['F2592'].value)*ws['F2677'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
        
        for item, items in zip(competitor, year):
            if item == 'Jusung Engineering' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.40))
                    
                    ws['F3229'] = total
                    
                    tab = ['F3239', 'F3243', 'F3245', 'F3246', 'F3247']
                                        
                    ratio = [(ws['F3154'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3158'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3160'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3161'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3162'].value/ws['F3144'].value)*ws['F3229'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
        
        for item, items in zip(competitor, year):
            if item == 'Lam Research' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F3781'] = total  
                    
                    tab = ['F3791', 'F3795', 'F3797', 'F3798', 'F3799', 'F3801', 'F3802', 'F3804', 'F3808', 'F3809', 'F3811', 'F3812', 'F3817', 'F3818', 'F3819', 'F3820', 'F3821']
                                        
                    ratio = [(ws['F3706'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3710'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3712'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3713'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3714'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3716'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3717'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3719'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3723'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3724'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3726'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3727'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3732'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3733'].value/ws['F3696'].value)*ws['F3781'].value,(ws['F3734'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3735'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3736'].value/ws['F3696'].value)*ws['F3781'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Orbotech' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F4307'] = total   
                    
                    tab = ['F4317', 'F4321', 'F4323', 'F4328', 'F4329', 'F4334', 'F4344', 'F4345', 'F4347', 'F4348', 'F4350', 'F4389', 'F4391']
                                        
                    ratio = [(ws['F4232'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4236'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4238'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4243'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4244'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4249'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4259'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4260'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4262'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4263'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4265'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4304'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4306'].value/ws['F4222'].value)*ws['F4307'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Tes' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.92))
                    
                    ws['F4859'] = total   
                    
                    tab = ['F4869', 'F4873', 'F4875', 'F4879', 'F4880', 'F4886', 'F4892', 'F4894']
                                        
                    ratio = [(ws['F4784'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4788'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4790'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4794'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4795'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4801'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4807'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4809'].value/ws['F4774'].value)*ws['F4859'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Tokyo Electron' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
           
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.17))
                    
                    ws['F5411'] = total  
                    
                    tab = ['F5420', 'F5421', 'F5425', 'F5426', 'F5427', 'F5428', 'F5430', 'F5431', 'F5432', 'F5433', 'F5434', 'F5438', 'F5440', 'F5441', 'F5442', 'F5443', 'F5444', 'F5445', 'F5446', 'F5448', 'F5449', 'F5450', 'F5451', 'F5456', 'F5459', 'F5460', 'F5493', 'F5495']
                                        
                    ratio = [(ws['F5335'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5336'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5340'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5341'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5342'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5343'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5345'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5346'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5347'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5348'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5349'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5353'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5355'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5356'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5357'].value/ws['F5326'].value)*ws['F5411'].value,(ws['F5358'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5359'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5360'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5361'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5363'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5364'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5365'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5366'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5371'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5374'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5375'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5408'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5410'].value/ws['F5326'].value)*ws['F5411'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Ultratech' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F5963'] = total   
                    
                    tab = ['F5964', 'F5966', 'F5970', 'F5973', 'F5977', 'F5979', 'F5980', 'F5981', 'F6008', 'F6009', 'F6018', 'F6024', 'F6026']
                                        
                    ratio = [(ws['F5879'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5881'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5886'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5888'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5892'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5894'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5895'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5896'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5923'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5924'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5933'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5939'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5941'].value/ws['F5878'].value)*ws['F5963'].value]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
        for item, items in zip(competitor, year):
            if item == 'Veeco' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                                                  
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F6515'] = total 
                    
                    tab = ['F6516', 'F6518', 'F6523', 'F6525', 'F6526', 'F6529', 'F6531', 'F6532', 'F6533', 'F6539', 'F6540', 'F6560', 'F6561', 'F6570', 'F6576', 'F6578']
                                        
                    ratio = [(ws['F6431'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6433'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6438'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6440'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6441'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6444'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6446'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6447'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6448'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6454'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6455'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6475'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6476'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6485'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6491'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6493'].value/ws['F6430'].value)*ws['F6515'].value ]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
        for item, items in zip(competitor, year):
            if item == 'Wonik IPS' and items == '2018': 
                if quarter_3 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    
                    quarter_4 = quarter_1+quarter_2-quarter_3
                    
                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = quarter_4 * ratio_cal
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F7067'] = total  
                    
                    tab = ['F7077', 'F7081', 'F7083', 'F7084', 'F7085', 'F7087', 'F7088']
                                        
                    ratio = [(ws['F6992'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6996'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6998'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6999'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7000'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7002'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7003'].value/ws['F6982'].value)*ws['F7067'].value ]
                    
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
        
    def quarter4(self):
        pre_quarter_1 = self.nameLine.text()
        quarter_1 = float(pre_quarter_1)
        pre_quarter_2 = self.nameLine1.text()
        quarter_2 = float(pre_quarter_2)
        pre_quarter_3 = self.nameLine2.text()
        quarter_3 = float(pre_quarter_3)
        pre_quarter_4 = self.nameLine3.text()
        quarter_4 = float(pre_quarter_4)
        pre_ratio = self.ratio.text()
        ratio = float(pre_ratio)
        
        for item, items in zip(competitor, year):
            if item == 'Aixtron' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F1021'] = total   
                    
                    tab = ['F1031', 'F1032', 'F1035', 'F1037', 'F1038', 'F1039', 'F1041']
                                        
                    ratio = [(ws['F946'].value/ws['F936'].value)*ws['F1021'].value, (ws['F947'].value/ws['F936'].value)*ws['F1021'].value, (ws['F950'].value/ws['F936'].value)*ws['F1021'].value, (ws['F952'].value/ws['F936'].value)*ws['F1021'].value, (ws['F953'].value/ws['F936'].value)*ws['F1021'].value, (ws['F954'].value/ws['F936'].value)*ws['F1021'].value, (ws['F956'].value/ws['F936'].value)*ws['F1021'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
        
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
              
                
        for item, items in zip(competitor, year):
            if item == 'ASM International' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.22))
                    
                    ws['F1573'] = total   
                    
                    tab = ['F1583', 'F1585', 'F1587', 'F1588', 'F1589', 'F1590', 'F1591', 'F1594', 'F1618', 'F1622']
                                        
                    ratio = [(ws['F1498'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1500'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1502'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1503'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1504'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1505'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1506'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1509'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1533'].value/ws['F1488'].value)*ws['F1573'].value, (ws['F1537'].value/ws['F1488'].value)*ws['F1573'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                
        for item, items in zip(competitor, year):
            if item == 'Eugene Technology' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.84))
                    
                    ws['F2125'] = total
                    
                    tab = ['F2135', 'F2137', 'F2139', 'F2141', 'F2142', 'F2143', 'F2145', 'F2170', 'F2173']
                                        
                    ratio = [(ws['F2050'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2052'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2054'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2056'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2057'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2058'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2060'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2085'].value/ws['F2040'].value)*ws['F2125'].value, (ws['F2088'].value/ws['F2040'].value)*ws['F2125'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                       
        for item, items in zip(competitor, year):
            if item == 'Hitachi Kokusai Electric' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                                        
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F2677'] = quart4   
                    
                    tab = ['F2687', 'F2691', 'F2692', 'F2704', 'F2713', 'F2722', 'F2723', 'F2725', 'F2726']
                                        
                    ratio = [(ws['F2602'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2606'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2607'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2619'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2628'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2637'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2638'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2640'].value/ws['F2592'].value)*ws['F2677'].value, (ws['F2641'].value/ws['F2592'].value)*ws['F2677'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
        for item, items in zip(competitor, year):
            if item == 'Jusung Engineering' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.40))
                    
                    ws['F3229'] = quart4   
                    
                    tab = ['F3239', 'F3243', 'F3245', 'F3246', 'F3247']
                                        
                    ratio = [(ws['F3154'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3158'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3160'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3161'].value/ws['F3144'].value)*ws['F3229'].value, (ws['F3162'].value/ws['F3144'].value)*ws['F3229'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)                   
                    
        for item, items in zip(competitor, year):
            if item == 'Lam Research' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F3781'] = total 
                    
                    tab = ['F3791', 'F3795', 'F3797', 'F3798', 'F3799', 'F3801', 'F3802', 'F3804', 'F3808', 'F3809', 'F3811', 'F3812', 'F3817', 'F3818', 'F3819', 'F3820', 'F3821']
                                        
                    ratio = [(ws['F3706'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3710'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3712'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3713'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3714'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3716'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3717'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3719'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3723'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3724'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3726'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3727'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3732'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3733'].value/ws['F3696'].value)*ws['F3781'].value,(ws['F3734'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3735'].value/ws['F3696'].value)*ws['F3781'].value, (ws['F3736'].value/ws['F3696'].value)*ws['F3781'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                       
        for item, items in zip(competitor, year):
            if item == 'Orbotech' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']

                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F4307'] = total  
                    
                    tab = ['F4317', 'F4321', 'F4323', 'F4328', 'F4329', 'F4334', 'F4344', 'F4345', 'F4347', 'F4348', 'F4350', 'F4389', 'F4391']
                                        
                    ratio = [(ws['F4232'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4236'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4238'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4243'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4244'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4249'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4259'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4260'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4262'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4263'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4265'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4304'].value/ws['F4222'].value)*ws['F4307'].value, (ws['F4306'].value/ws['F4222'].value)*ws['F4307'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)                    
                    
        for item, items in zip(competitor, year):
            if item == 'Tes' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                              
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.92))
                    
                    ws['F4859'] = total 
                    
                    tab = ['F4869', 'F4873', 'F4875', 'F4879', 'F4880', 'F4886', 'F4892', 'F4894']
                                        
                    ratio = [(ws['F4784'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4788'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4790'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4794'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4795'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4801'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4807'].value/ws['F4774'].value)*ws['F4859'].value, (ws['F4809'].value/ws['F4774'].value)*ws['F4859'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)
                    
                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
        for item, items in zip(competitor, year):
            if item == 'Tokyo Electron' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.17))
                    
                    ws['F5411'] = total   
                    
                    tab = ['F5420', 'F5421', 'F5425', 'F5426', 'F5427', 'F5428', 'F5430', 'F5431', 'F5432', 'F5433', 'F5434', 'F5438', 'F5440', 'F5441', 'F5442', 'F5443', 'F5444', 'F5445', 'F5446', 'F5448', 'F5449', 'F5450', 'F5451', 'F5456', 'F5459', 'F5460', 'F5493', 'F5495']
                                        
                    ratio = [(ws['F5335'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5336'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5340'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5341'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5342'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5343'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5345'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5346'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5347'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5348'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5349'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5353'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5355'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5356'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5357'].value/ws['F5326'].value)*ws['F5411'].value,(ws['F5358'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5359'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5360'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5361'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5363'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5364'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5365'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5366'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5371'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5374'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5375'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5408'].value/ws['F5326'].value)*ws['F5411'].value, (ws['F5410'].value/ws['F5326'].value)*ws['F5411'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                    
        for item, items in zip(competitor, year):
            if item == 'Ultratech' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F5963'] = total 
                    
                    tab = ['F5964', 'F5966', 'F5970', 'F5973', 'F5977', 'F5979', 'F5980', 'F5981', 'F6008', 'F6009', 'F6018', 'F6024', 'F6026']
                                        
                    ratio = [(ws['F5879'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5881'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5886'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5888'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5892'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5894'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5895'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5896'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5923'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5924'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5933'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5939'].value/ws['F5878'].value)*ws['F5963'].value, (ws['F5941'].value/ws['F5878'].value)*ws['F5963'].value]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    
                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
        for item, items in zip(competitor, year):
            if item == 'Veeco' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.148))
                    
                    ws['F6515'] = total   
                    
                    tab = ['F6516', 'F6518', 'F6523', 'F6525', 'F6526', 'F6529', 'F6531', 'F6532', 'F6533', 'F6539', 'F6540', 'F6560', 'F6561', 'F6570', 'F6576', 'F6578']
                                        
                    ratio = [(ws['F6431'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6433'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6438'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6440'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6441'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6444'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6446'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6447'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6448'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6454'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6455'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6475'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6476'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6485'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6491'].value/ws['F6430'].value)*ws['F6515'].value, (ws['F6493'].value/ws['F6430'].value)*ws['F6515'].value ]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
        for item, items in zip(competitor, year):
            if item == 'Wonik IPS' and items == '2018': 
                if quarter_4 != "":
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb = load_workbook(filepath)
                    ws = wb['Revised']
                    
                    ratio_cal = float((100-ratio)/ratio)
                    quarters = ((quarter_1+quarter_2)/(quarter_3+quarter_4))
                    quarter_3_4 = ((quarter_3+quarter_4)*(quarters)/2)

                    quart1 = float(quarter_1) 
                    quart2 = float(quarter_2)
                    quart3 = float(quarter_3)
                    quart4 = float(quarter_4)
                    
                    total = float((quart1+quart2+ quart3+ quart4)-((quart1+quart2+ quart3+ quart4)*.34))
                    
                    ws['F7067'] = total  
                    
                    tab = ['F7077', 'F7081', 'F7083', 'F7084', 'F7085', 'F7087', 'F7088']
                                        
                    ratio = [(ws['F6992'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6996'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6998'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F6999'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7000'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7002'].value/ws['F6982'].value)*ws['F7067'].value, (ws['F7003'].value/ws['F6982'].value)*ws['F7067'].value ]
                                      
                    for cell, num in zip(tab, ratio):
                        ws[cell] = num 
                    
                    wb.save(filepath)
                    
                    filepath = 'Competitor-Flat-File.xlsx'
                    wb1 = load_workbook(filepath)
                    ws1 = wb1['QoQ']
                    
                    ws1['B2'] = quart1
                    ws1['B3'] = quart2
                    ws1['B4'] = quart3
                    ws1['B5'] = quart4
                    
                    wb1.save(filepath)

                    self.verified1.setText("%s" % quart1)
                    self.verified2.setText("%s" % quart2)
                    self.verified3.setText("%s" % quart3)
                    self.verified4.setText("%s" % quart4)
                        
    def graph_ratio(self):
        name1 = (self.nameLine4.text())
        name2 = (self.nameLine5.text())
        pre_ratio = self.ratio.text()
        ratio = float(pre_ratio)
                
        if name1 == 'Aixtron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Aixtron QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'ASM International' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("ASM International QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Eugene Technology' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Eugene Technology QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Hitachi Kokusai Electric' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Hitachi Kokusai Electric QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Jusung Engineering' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Jusung Engineering QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
        
        if name1 == 'Lam Research' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Lam Research QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Orbotech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Orbotech QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Tes' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Tes QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Tokyo Electron' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Tokyo Electron QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Ultratech' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Ultratech QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Veeco' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Veeco QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
            
        if name1 == 'Wonik IPS' and name2 == '2018':
            filepath = 'Competitor-Flat-File.xlsx'
            wb = load_workbook(filepath)
            ws = wb['QoQ']
            
            qn1 = ws['A2'].value
            qn2 = ws['A3'].value
            qn3 = ws['A4'].value
            qn4 = ws['A5'].value
            
            q1 = ws['B2'].value
            q2 = ws['B3'].value 
            q3 = ws['B4'].value
            q4 = ws['B5'].value
            
            q5 = q1+q2+q3+q4
            
            fig, ax = plt.subplots(figsize = (7,8))
            plt.subplots_adjust(left=0.25, bottom=0.25)
            ratio = mpatches.Patch(label = 'Ratio Q1 & Q2:%s'%ratio, color = 'cyan')
            plt.bar(qn1, q1, lw=2, label = 'Q1:%s'%q1,color='blue')
            plt.bar(qn2, q2, lw=2, label = 'Q2:%s'%q2, color='blue')
            plt.bar(qn3, q3, lw=2, label = 'Q3:%s'%q3,color='green')
            plt.bar(qn4, q4, lw=2, label = 'Q4:%s'%q4,color='green')
            
            plt.title("Wonik IPS QoQ")
            first_legend = plt.legend(handles =[ratio], loc='upper left', prop={'size':7}, bbox_to_anchor=(1,1))

            ax = plt.gca().add_artist(first_legend)
            plt.legend(loc=6, prop={'size':7}, bbox_to_anchor=(1,0.8))
            plt.tight_layout(pad=7) 
            plt.ylabel('Yearly Revenue')
            plt.show()
                    
    def submitContact(self): 
        self.close()
              
if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = First()
    main.show()
    app.exec_()